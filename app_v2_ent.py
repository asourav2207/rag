import streamlit as st
import os
import pandas as pd
from datetime import datetime
from sentence_transformers import SentenceTransformer
import numpy as np
import faiss
import requests
import json
from sqlalchemy import create_engine, Column, Integer, String, Text, DateTime, text
from sqlalchemy.orm import declarative_base, sessionmaker
import base64
from io import BytesIO
from openpyxl import load_workbook
import concurrent.futures
import re
from urllib.parse import quote_plus # Import this

# ---- App Configuration ----
st.set_page_config(page_title="Redshift Assistant", layout="wide")
st.title("🧑‍💻 Redshift Multipurpose Assistant")

# ---- GitHub Integration Config ----
GITHUB_TOKEN = st.secrets.get("github_token", os.environ.get("GITHUB_TOKEN", ""))
GITHUB_REPO = st.secrets.get("github_repo", "asourav2207/rag")
GITHUB_BRANCH = st.secrets.get("github_branch", "main")
GITHUB_EXCEL_PATH = st.secrets.get("github_excel_path", "excel_files")

# ---- Redshift & Database Setup ----
Base = declarative_base()
DB_PATH = "sqlite:///chat_history.db"

# ---- SQLAlchemy ORM Models ----
class ChatSession(Base):
    __tablename__ = 'chat_session'
    id = Column(Integer, primary_key=True)
    name = Column(String, default="Untitled Chat")
    mode = Column(String, default="Documentation")
    created_at = Column(DateTime, default=datetime.now)

class ChatHistory(Base):
    __tablename__ = 'chat_history'
    id = Column(Integer, primary_key=True)
    question = Column(Text)
    answer = Column(Text)
    context = Column(Text) # This will now store JSON string for rich context display
    sql_query = Column(Text)
    timestamp = Column(DateTime, default=datetime.now)
    feedback = Column(String)
    session_id = Column(Integer)

engine = create_engine(DB_PATH)
Base.metadata.create_all(engine)
SessionLocal = sessionmaker(bind=engine)

# ---- Redshift Functions ----
@st.cache_resource(show_spinner="Connecting to Redshift...")
def get_redshift_engine():
    try:
        creds = st.secrets["redshift"]
        
        # URL-encode the password to handle special characters
        encoded_password = quote_plus(creds['password'])

        conn_string = (
            f"postgresql+psycopg2://{creds['user']}:{encoded_password}@"
            f"{creds['host']}:{creds['port']}/{creds['database']}"
        )
        engine = create_engine(conn_string, connect_args={'connect_timeout': 10})
        with engine.connect() as conn:
            st.sidebar.success("Redshift connection successful!")
        return engine
    except Exception as e:
        st.sidebar.error(f"Redshift connection failed: {e}")
        return None

@st.cache_data(ttl=3600, show_spinner="Fetching Redshift schema...")
def get_redshift_schema(_engine):
    if _engine is None: return "Redshift connection not available."
    schema_query = """
    SELECT table_schema, table_name, column_name, data_type
    FROM information_schema.columns
    WHERE table_schema NOT IN ('information_schema', 'pg_catalog')
    ORDER BY table_schema, table_name, ordinal_position;
    """
    try:
        with _engine.connect() as conn:
            result = conn.execute(text(schema_query))
            schema_df = pd.DataFrame(result.fetchall(), columns=result.keys())
        schema_str = ""
        for (schema, table), group in schema_df.groupby(['table_schema', 'table_name']):
            schema_str += f"Table: {schema}.{table}\n"
            for _, row in group.iterrows():
                schema_str += f"  - {row['column_name']} ({row['data_type']})\n"
            schema_str += "\n"
        return schema_str
    except Exception as e:
        st.error(f"Failed to fetch Redshift schema: {e}")
        return None

def run_redshift_query(sql, engine):
    if engine is None:
        st.error("Cannot run query, Redshift connection is not available.")
        return None
    try:
        with engine.connect() as conn:
            result = conn.execute(text(sql))
            if result.returns_rows:
                return pd.DataFrame(result.fetchall(), columns=result.keys())
            else:
                st.success("Query executed successfully.")
                return pd.DataFrame([{"status": "Success", "rows_affected": result.rowcount}])
    except Exception as e:
        st.error(f"Redshift query error: {e}")
        return None

def extract_sql_from_answer(answer):
    match = re.search(r'```sql\s*(.*?)\s*```', answer, re.DOTALL | re.IGNORECASE)
    return match.group(1).strip() if match else None

# ---- GitHub, Excel, & Changelog Functions ----
def github_headers():
    return {"Authorization": f"token {GITHUB_TOKEN}", "Accept": "application/vnd.github.v3+json"}

def list_github_excels():
    url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{GITHUB_EXCEL_PATH}?ref={GITHUB_BRANCH}"
    try:
        r = requests.get(url, headers=github_headers(), timeout=10)
        r.raise_for_status()
        return [f for f in r.json() if f["name"].endswith((".xlsx", ".xls"))]
    except Exception:
        return []

def download_excel_from_github(file_info):
    try:
        r = requests.get(file_info["download_url"], timeout=10)
        r.raise_for_status()
        # FIX: Wrap r.content in BytesIO for pandas compatibility
        return pd.read_excel(BytesIO(r.content), engine="openpyxl")
    except Exception as e:
        st.error(f"Error downloading {file_info['name']}: {e}")
        return None

def generate_changelog(existing_df, new_df):
    diffs = []
    # Simple diff: checks for changes in values, row by row
    for i in range(max(len(existing_df), len(new_df))):
        if i >= len(existing_df):
            diffs.append(f"Row {i+1} added: {new_df.iloc[i].to_dict()}")
            continue
        if i >= len(new_df):
            diffs.append(f"Row {i+1} removed: {existing_df.iloc[i].to_dict()}")
            continue
        
        old_row, new_row = existing_df.iloc[i], new_df.iloc[i]
        if not old_row.equals(new_row):
            for col in new_df.columns:
                if col in old_row and str(old_row[col]) != str(new_row[col]): # Cast to str for comparison robustness
                    diffs.append(f"Row {i+1}, Column '{col}': changed from '{old_row[col]}' to '{new_row[col]}'")
    return diffs

def inject_changelog_to_excel(uploaded_file, changelog_lines, user_id):
    uploaded_file.seek(0)
    wb = load_workbook(uploaded_file)
    ws = wb.create_sheet("Changelog") if "Changelog" not in wb.sheetnames else wb["Changelog"]
    if ws.max_row == 1: # Add header if sheet is new/empty (or only has 1 row which could be just header)
        if ws.cell(row=1, column=1).value not in ["Timestamp", "user", "User", "Change Description"]: # Basic check for header
            ws.append(["Timestamp", "User", "Change Description"])
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for line in changelog_lines:
        ws.append([timestamp, user_id, line])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def upload_excel_to_github(file, filename):
    url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{GITHUB_EXCEL_PATH}/{filename}"
    try:
        get_resp = requests.get(url, headers=github_headers(), timeout=10)
        sha = get_resp.json().get("sha") if get_resp.status_code == 200 else None
        file.seek(0)
        content = base64.b64encode(file.read()).decode()
        data = {"message": f"Update {filename}", "content": content, "branch": GITHUB_BRANCH}
        if sha: data["sha"] = sha
        put_resp = requests.put(url, headers=github_headers(), data=json.dumps(data), timeout=15)
        put_resp.raise_for_status()
        return True
    except Exception as e:
        st.error(f"GitHub upload failed for {filename}: {e}")
        return False

# ---- RAG (Documentation Mode) Functions ----
@st.cache_resource(show_spinner="Loading documentation model and data...")
def load_and_prepare_rag_data():
    row_docs, field_docs, row_metadata, field_metadata = [], [], [], []
    excel_files = list_github_excels()
    if not excel_files: st.warning("No Excel documentation files found in GitHub.")
    for file_info in excel_files:
        df = download_excel_from_github(file_info)
        if df is None: continue
        table_name = os.path.splitext(file_info["name"])[0]
        
        # Check if the Excel file has a "Changelog" sheet and exclude it from documentation data
        try:
            workbook = load_workbook(BytesIO(requests.get(file_info["download_url"], timeout=10).content))
            if "Changelog" in workbook.sheetnames:
                st.info(f"Skipping 'Changelog' sheet in {file_info['name']} for RAG processing.")
                # We'll read the first sheet as the primary data sheet
                df = pd.read_excel(BytesIO(requests.get(file_info["download_url"], timeout=10).content), sheet_name=workbook.sheetnames[0], engine="openpyxl")
                if "Changelog" in df.columns: # If changelog column is still there, remove it
                     df = df.drop(columns=["Changelog"], errors='ignore')
        except Exception as e:
            st.warning(f"Could not check sheets for {file_info['name']}: {e}. Proceeding with default sheet.")

        for idx, row in df.iterrows():
            row_summary = [f"{col.strip()}: {str(row.get(col, '')).strip()}" for col in df.columns if str(row.get(col, '')).strip()]
            if row_summary:
                row_docs.append(f"Table: {table_name}, Row: {idx}\n" + "\n".join(row_summary))
                # Store full row data for richer context display
                row_metadata.append({'table': table_name, 'row_index': idx, 'content': "\n".join(row_summary), 'original_row_data': row.to_dict()})

    if not row_docs: return None, None, [], [], None, [], [] # Ensure to return appropriate empty values
    model = SentenceTransformer('all-MiniLM-L6-v2')
    dim = model.get_sentence_embedding_dimension()
    row_index = faiss.IndexFlatL2(dim)
    row_embeddings = model.encode(row_docs, show_progress_bar=False)
    row_index.add(np.array(row_embeddings))
    return model, row_index, row_docs, row_metadata, None, [], [] # field_index, field_docs, field_metadata are not used here

def retrieve_relevant_docs(query, model, row_index, row_docs, row_metadata, k=5):
    if not row_docs: return []
    
    # NEW: Query Expansion/Rewriting for better embedding match
    query_expansion_prompt = f"""Given the user query: "{query}"
    Generate 2-3 alternative phrasings or related terms that could be used to search documentation.
    Example:
    Query: "payment logic"
    Expanded: "how payments are calculated", "payment rules", "payment process"
    
    Expanded:"""
    
    # Temporarily using call_ollama, ideally a smaller, faster model or a dedicated API for query expansion.
    # Note: If Ollama is not running, this will return an error message string, so handle it.
    expanded_query_raw = call_ollama(query_expansion_prompt, timeout=10) 
    
    expanded_queries = []
    if not expanded_query_raw.startswith("Error contacting Ollama"): # Check if Ollama call was successful
        expanded_queries = [q.strip() for q in expanded_query_raw.split(',') if q.strip()]

    search_queries = [query] + expanded_queries[:2] # Use original + up to 2 expanded queries

    all_matches = []
    for q in search_queries:
        query_emb = model.encode([q])
        distances, indices = row_index.search(query_emb, k)
        # Filter by a distance threshold to only include truly relevant docs
        # The threshold (e.g., 1.0 or 1.5) depends on your embedding model and data
        for i, dist in zip(indices[0], distances[0]):
            if dist < 1.0: # Adjust threshold as needed
                all_matches.append((row_docs[i], row_metadata[i], dist)) # Store distance too for sorting

    # Sort matches by distance (lower is better) and remove duplicates
    all_matches.sort(key=lambda x: x[2])
    unique_matches = []
    seen_content = set()
    for doc, meta, dist in all_matches:
        if doc not in seen_content:
            unique_matches.append((doc, meta))
            seen_content.add(doc)
        if len(unique_matches) >= k: # Limit to top k unique matches
            break
            
    return unique_matches

# ---- LLM & Chat Functions ----
def call_ollama(prompt, timeout=90):
    try:
        response = requests.post("http://127.0.0.1:11434/api/generate",
            json={"model": "llama3.2", "prompt": prompt, "stream": False}, timeout=timeout)
        response.raise_for_status() # Corrected: raise_for_status()
        return response.json().get("response", "").strip()
    except requests.exceptions.RequestException as e:
        return f"Error contacting Ollama: {e}"

def get_rag_answer(user_query, rag_data, history):
    model, row_index, row_docs, row_metadata, _, _, _ = rag_data
    if not model or not row_index or not row_docs or not row_metadata:
        return "Documentation data not loaded or indexed. Please ensure Excel files are uploaded and data is available.", "", None

    # Handle ambiguous queries
    ambiguous_terms = ["this field", "it", "the field", "that field"]
    if any(term in user_query.lower() for term in ambiguous_terms) and st.session_state.get("last_referenced_field"):
        user_query += f" (referring to {st.session_state.last_referenced_field})"

    matches = retrieve_relevant_docs(user_query, model, row_index, row_docs, row_metadata)
    
    # Store last referenced field for context
    if matches:
        # Try to extract a field from the first match for follow-up context
        first_match_content = matches[0][1].get('content', '')
        m = re.search(r"([a-zA-Z0-9_]+)\s*:", first_match_content)
        if m: st.session_state["last_referenced_field"] = m.group(1)

    # Prepare context for LLM (still plain string, taking top 3 for brevity in prompt)
    context_for_llm = "\n---\n".join([doc for doc, meta in matches][:3])

    conversation_history = "".join([f"User: {h.question}\nAssistant: {h.answer}\n" for h in history[-2:] if h.answer != "Thinking..."]) # Exclude current "Thinking..." entry
    
    # Refined RAG Prompt for LLM
    prompt = f"""You are a helpful and precise documentation assistant. Your primary goal is to answer the user's question STRICTLY based on the provided "Documentation Context".
If the answer is not directly available or cannot be reasonably inferred from the context, you MUST state: "I cannot answer this question based on the provided documentation."
Do NOT use any outside knowledge or make assumptions. Be concise and direct.

Conversation History:
{conversation_history}

Documentation Context:
{context_for_llm}

User Query: {user_query}

Answer:"""
    answer = call_ollama(prompt)
    
    # Pass the full matches object as context to save, so it can be rendered later.
    # We'll JSON dump only the metadata for storage, not the full document string.
    return answer, json.dumps([m[1] for m in matches]), None

def get_sql_answer(user_query, schema_str, history):
    conversation_history = "".join([f"User: {h.question}\nAssistant:\n{h.answer}\n" for h in history[-3:] if h.answer != "Thinking..."])
    prompt = f"""You are an expert Redshift SQL analyst. Convert the user's question into a correct Redshift SQL query using the provided schema and history.
Conversation History:
{conversation_history}
Redshift Schema:
{schema_str}
User Question: {user_query}
Generate ONLY the SQL query in a single SQL code block. Example: ```sql\nSELECT * FROM my_table;\n```"""
    response = call_ollama(prompt)
    sql_query = extract_sql_from_answer(response)
    answer = "I have translated your request into the SQL query below. You can review, edit, and run it."
    if not sql_query: answer = "I was unable to generate a SQL query. Please try rephrasing."
    return answer, schema_str, sql_query # Store schema as context for SQL mode

def generate_followup_suggestions_async(user_prompt, system_response, context):
    def _llm_suggest():
        context_summary = ""
        try:
            if context and context.strip().startswith('['): # Heuristic for JSON context
                parsed_context = json.loads(context)
                if parsed_context:
                    context_summary = "Relevant Tables/Fields from Documentation:\n" + "\n".join(
                        [f"- {m.get('table', '')} (Row {m.get('row_index', '')})" for m in parsed_context[:2]]
                    )
            else: # Likely SQL schema context or simple text
                context_summary = f"Context snippet: {str(context)[:200]}..."
        except (json.JSONDecodeError, TypeError):
            context_summary = f"Context snippet: {str(context)[:200]}..."

        prompt = f"""Given the user question, system answer, and relevant context, generate 2-3 concise, direct follow-up questions a user might ask next. Phrase each as a question.
User question: {user_prompt}
System answer: {system_response}
{context_summary}
List the follow-up questions as bullet points."""
        
        # Check if Ollama is accessible before calling it
        suggestions = call_ollama(prompt, timeout=20)
        if suggestions.startswith("Error contacting Ollama"):
            return ["(Suggestions: Ollama not reachable)"]
        
        return [line.lstrip('-• ').strip() for line in suggestions.split('\n') if line.strip()]
    
    with concurrent.futures.ThreadPoolExecutor() as executor:
        future = executor.submit(_llm_suggest)
        try: 
            return future.result(timeout=25)
        except concurrent.futures.TimeoutError:
            return ["(Suggestions timed out)"]
        except Exception as e:
            return [f"(Error generating suggestions: {e})"]

def save_chat(question, answer, context, sql, session_id):
    db = SessionLocal()
    entry = ChatHistory(question=question, answer=answer, context=context, sql_query=sql, session_id=session_id)
    db.add(entry); db.commit(); db.close()

def get_chat_history(session_id):
    db = SessionLocal()
    entries = db.query(ChatHistory).filter_by(session_id=session_id).order_by(ChatHistory.timestamp).all()
    db.close()
    return entries

# Initialize session state variables
if 'selected_session_id' not in st.session_state:
    st.session_state.selected_session_id = None
if 'history' not in st.session_state:
    st.session_state.history = []
if 'user_input' not in st.session_state:
    st.session_state.user_input = ""
if 'rerun_triggered_by_suggestion' not in st.session_state:
    st.session_state.rerun_triggered_by_suggestion = False

# ---- Main App UI ----
st.sidebar.title("Controls & Sessions")
app_mode = st.sidebar.toggle("Enable Redshift Query Assistant", value=False)
mode_label = "Redshift Query" if app_mode else "Documentation"
st.sidebar.markdown(f"**Current Mode:** {mode_label}")

with st.sidebar.form("new_session_form"):
    new_chat_name = st.text_input("New Chat Session Name")
    if st.form_submit_button("Start New Chat") and new_chat_name:
        db = SessionLocal()
        new_session = ChatSession(name=new_chat_name, mode=mode_label)
        db.add(new_session); db.commit()
        st.session_state.selected_session_id = new_session.id
        db.close()
        st.session_state.history = [] # Clear history for new session
        st.session_state.user_input = "" # Clear any pending input
        st.session_state.rerun_triggered_by_suggestion = False # Reset
        st.rerun()

db = SessionLocal()
sessions = db.query(ChatSession).order_by(ChatSession.created_at.desc()).all()
db.close()
session_options = {f"{s.name} ({s.mode})": s.id for s in sessions}

# Handle initial session selection or session change
if not st.session_state.selected_session_id and sessions:
    st.session_state.selected_session_id = sessions[0].id

if st.session_state.selected_session_id:
    # Find the key for the currently selected ID to set default in selectbox
    current_session_key = next((key for key, value in session_options.items() if value == st.session_state.selected_session_id), None)
    
    # If the current_session_key is not found (e.g., session was deleted externally),
    # or if there are no sessions at all, handle gracefully.
    options_list = list(session_options.keys())
    selected_index = 0
    if current_session_key and current_session_key in options_list:
        selected_index = options_list.index(current_session_key)

    selected_session_name = st.sidebar.selectbox("Select Chat Session", 
                                                options=options_list, 
                                                index=selected_index)
    
    newly_selected_id = session_options[selected_session_name]
    if newly_selected_id != st.session_state.selected_session_id:
        st.session_state.selected_session_id = newly_selected_id
        # Reload history only if the session actually changed
        st.session_state.history = get_chat_history(st.session_state.selected_session_id)
        st.session_state.user_input = "" # Clear input on session change
        st.session_state.rerun_triggered_by_suggestion = False # Reset
        st.rerun() # Rerun to refresh chat display with new session's history
elif not sessions:
    st.sidebar.info("No chat sessions found. Start a new one above!")
    st.info("Start a new chat session using the sidebar to begin.")


# --- Main Content Area ---
redshift_engine = None
schema_info = None
rag_data = None

if app_mode: # Redshift Query Mode
    st.header("Redshift Query Assistant")
    st.info("Ask a question in natural language to query your Redshift database.")
    redshift_engine = get_redshift_engine()
    schema_info = get_redshift_schema(redshift_engine) if redshift_engine else None
    if not schema_info or "not available" in schema_info:
        st.error("Cannot proceed without Redshift schema. Check connection details in secrets.")
    else:
        with st.expander("View Detected Redshift Schema"):
            st.code(schema_info, language='text')
else: # Documentation Mode
    st.header("Documentation Assistant")
    st.info("Ask questions based on the documentation in your GitHub Excel files.")
    rag_data = load_and_prepare_rag_data()
    if not rag_data or not rag_data[0]: # Check if model loaded successfully (model is the first element)
        st.warning("Documentation data could not be loaded. Ensure Excel files are present and accessible in GitHub.")

    with st.expander("Upload & Manage Documentation"):
        enable_changelog = st.toggle("Enable Changelog Tracking", value=True)
        uploaded_file = st.file_uploader("Upload Excel file to GitHub", type=["xlsx", "xls"])
        
        if uploaded_file:
            st.session_state.last_uploaded_filename = uploaded_file.name # Store for consistent key
            existing_df = None
            excel_files = list_github_excels()
            existing_file_info = [f for f in excel_files if f["name"] == uploaded_file.name]
            
            if existing_file_info:
                existing_df = download_excel_from_github(existing_file_info[0])

            file_to_upload = uploaded_file
            changelog_generated = False
            
            if enable_changelog and existing_df is not None:
                user_id = st.text_input("Enter your User ID for changelog", key=f"user_id_{st.session_state.last_uploaded_filename}")
                if user_id:
                    try:
                        new_df = pd.read_excel(uploaded_file, engine="openpyxl")
                        changelog = generate_changelog(existing_df, new_df)
                        if changelog:
                            file_to_upload = inject_changelog_to_excel(uploaded_file, changelog, user_id)
                            st.info("Changelog generated and added to the file.")
                            changelog_generated = True
                        else:
                            st.info("No significant changes detected in the uploaded file compared to existing version.")
                    except Exception as e:
                        st.error(f"Error processing changelog: {e}")
            
            if st.button(f"Upload {uploaded_file.name}"):
                if upload_excel_to_github(file_to_upload, uploaded_file.name):
                    st.success(f"✅ {uploaded_file.name} uploaded. Refreshing documentation data...")
                    st.cache_resource.clear() # Clear cache to force reload RAG data
                    st.rerun() # Rerun to load new data and suggestions (and re-init app)

            if uploaded_file and rag_data and rag_data[0]: # Show smart suggestions only after upload button clicked and model ready
                st.subheader("💡 Smart Suggestions (based on uploaded file)")
                try:
                    new_df = pd.read_excel(uploaded_file, engine="openpyxl") # Read again for column names
                    cols = new_df.columns.str.lower()
                    
                    # Check for common columns and suggest questions
                    suggested_buttons = []
                    if "amount" in cols or "payment" in cols:
                        suggested_buttons.append("What is the logic for payment amount?")
                    if "date" in cols:
                        suggested_buttons.append("Are all date fields in the same format?")
                    
                    table_name_from_upload = uploaded_file.name.split('.')[0]
                    suggested_buttons.append(f"Generate sample documentation query for {table_name_from_upload}")
                    
                    for sugg in suggested_buttons:
                        if st.button(sugg, key=f"suggestion_after_upload_{sugg}"):
                            st.session_state.user_input = sugg
                            st.session_state.rerun_triggered_by_suggestion = True # Set flag
                            st.rerun()
                except Exception as e:
                    st.warning(f"Could not generate smart suggestions for uploaded file: {e}")


# --- Chat Interface ---
# Display messages
for i, entry in enumerate(st.session_state.get('history', [])):
    with st.chat_message("user"):
        st.markdown(entry.question)
    with st.chat_message("assistant"):
        st.markdown(entry.answer)
        if app_mode and entry.sql_query: # Only show SQL query in Redshift mode
            st.code(entry.sql_query, language="sql")
        
        # Display rich context in Documentation mode
        if not app_mode and entry.context:
            try:
                retrieved_metadata = json.loads(entry.context)
                if retrieved_metadata: # Only show expander if there's actual parsed context
                    with st.expander("📚 Retrieved Documentation Context"):
                        for meta in retrieved_metadata:
                            st.markdown(f"**Table:** `{meta.get('table', 'N/A')}` | **Row Index:** `{meta.get('row_index', 'N/A')}`")
                            # Display original_row_data as a DataFrame for better tabular view
                            if meta.get('original_row_data'):
                                # Convert the single row dictionary to a DataFrame for display
                                df_row = pd.DataFrame([meta['original_row_data']])
                                st.dataframe(df_row.T.rename(columns={0: 'Value'})) # Transpose for better readability of single row
                            else:
                                st.code(meta.get('content', 'No content available'), language='text')
                            st.markdown("---") # Separator between rows
                else:
                     with st.expander("📚 No Specific Documentation Context Retrieved"):
                        st.markdown("The assistant did not find specific documentation rows relevant to this query. It answered based on its general understanding or broad context.")

            except json.JSONDecodeError:
                # Fallback for old entries or if context is not valid JSON
                with st.expander("📚 Raw Retrieved Context (Legacy)"):
                    st.markdown(entry.context)
            except Exception as e:
                st.error(f"Error displaying context: {e}")
                with st.expander("📚 Error Displaying Context"):
                    st.markdown(f"Could not parse context: {entry.context[:100]}...")


        # Follow-up suggestions for the last message
        if i == len(st.session_state.history) - 1 and entry.answer != "Thinking...": # Only show suggestions for the final answer
            with st.expander("🔎 Follow-up Suggestions"):
                # Pass the raw context string to the async function
                suggestions = generate_followup_suggestions_async(entry.question, entry.answer, entry.context)
                for sugg in suggestions:
                    # When a suggestion button is clicked, set user_input and trigger rerun
                    # Using a unique key and a callback to ensure state update before rerun
                    if st.button(f"👉 {sugg}", key=f"sugg_{entry.id}_{sugg}"):
                        st.session_state.user_input = sugg
                        st.session_state.rerun_triggered_by_suggestion = True # Set flag
                        st.rerun()

# --- Input Box Logic ---

# Check if a rerun was triggered by a suggestion click. If so, immediately process the stored input.
if st.session_state.rerun_triggered_by_suggestion:
    st.session_state.rerun_triggered_by_suggestion = False # Reset the flag
    # If user_input is set by a suggestion, this block will be executed immediately on rerun
    if st.session_state.user_input and st.session_state.selected_session_id:
        question_to_process = st.session_state.user_input
        # Append a placeholder "Thinking..." message immediately
        st.session_state.history.append(ChatHistory(question=question_to_process, answer="Thinking...", session_id=st.session_state.selected_session_id))
        st.session_state.user_input = "" # Clear input for next user entry
        st.rerun() # Rerun to show "Thinking..." and trigger processing

# The chat_input widget's value will be what the user types OR what we set via st.session_state.user_input
# This allows suggestions to pre-fill the box and trigger a new message.
user_input_from_chatbox = st.chat_input(
    f"Ask a question... (Mode: {mode_label})", 
    key="chat_box"
)

# Check if there's new input from the chatbox (user typed something)
if user_input_from_chatbox and user_input_from_chatbox != st.session_state.user_input:
    st.session_state.user_input = user_input_from_chatbox
    st.rerun() # Rerun to trigger the processing logic

# Process the user input if it's new and we have a session AND not already processing
if st.session_state.user_input and st.session_state.selected_session_id:
    # Check if the last message is still "Thinking..." or if this is a repeat of the last question
    last_q_in_history = st.session_state.history[-1].question if st.session_state.history else None
    last_a_in_history = st.session_state.history[-1].answer if st.session_state.history else None

    # Only process if:
    # 1. The input is not empty
    # 2. It's a new question (not matching the last question if the last answer is not "Thinking..." AND there's a question)
    #    The `and last_q_in_history` part handles the very first question when history is empty.
    # 3. The last answer is not currently "Thinking..." (to prevent re-triggering while a response is being generated)
    if st.session_state.user_input.strip() != "" and \
       (st.session_state.user_input != last_q_in_history or (last_a_in_history and last_a_in_history != "Thinking...")):
        
        question_to_process = st.session_state.user_input
        # Append a placeholder "Thinking..." message immediately
        st.session_state.history.append(ChatHistory(question=question_to_process, answer="Thinking...", session_id=st.session_state.selected_session_id))
        st.session_state.user_input = "" # Clear input for next user entry
        st.rerun() # Rerun to show "Thinking..." and trigger processing

# Background processing for the "Thinking..." message
# This block runs AFTER the app reruns (either by user typing, suggestion click, or previous processing finishing)
if st.session_state.get('history') and st.session_state.history[-1].answer == "Thinking...":
    last_entry_placeholder = st.session_state.history[-1]
    with st.spinner("🧠 Thinking..."):
        answer, context, sql = None, None, None # Initialize
        if app_mode: # Redshift Query Mode
            if redshift_engine and schema_info:
                answer, context, sql = get_sql_answer(last_entry_placeholder.question, schema_info, st.session_state.history[:-1])
            else:
                answer = "Redshift connection or schema not available to generate query."
                context = ""
                sql = ""
        else: # Documentation Mode
            if rag_data and rag_data[0]: # Check if model is loaded (rag_data[0] is the model)
                answer, context, sql = get_rag_answer(last_entry_placeholder.question, rag_data, st.session_state.history[:-1])
            else:
                answer = "Documentation model not loaded. Please upload Excel files or check configuration."
                context = ""
                sql = ""
        
        # Update the last entry in history with the actual response
        last_entry_placeholder.answer = answer
        last_entry_placeholder.context = context
        last_entry_placeholder.sql_query = sql
        
        # Save the updated entry to the database
        save_chat(last_entry_placeholder.question, answer, context, sql, st.session_state.selected_session_id)
        
        # Re-fetch history to ensure the updated entry is reflected and session is consistent
        st.session_state.history = get_chat_history(st.session_state.selected_session_id)
        st.rerun() # Rerun to display the final answer and clear spinner

# --- SQL Execution Box (only visible in Redshift Query Mode) ---
if st.session_state.get('history') and app_mode:
    last_entry = st.session_state.history[-1]
    # Only show SQL execution if it's the last response and it generated a SQL query
    if last_entry.sql_query and last_entry.answer != "Thinking...":
        st.markdown("---")
        st.subheader("SQL Execution")
        # Ensure unique key for the text area and button
        edited_sql = st.text_area("Review and edit SQL:", value=last_entry.sql_query, height=150, key=f"sql_edit_{last_entry.id}_{last_entry.timestamp}")
        if st.button("▶️ Run Query on Redshift", key=f"run_sql_{last_entry.id}_{last_entry.timestamp}"):
            with st.spinner("Running query..."):
                results_df = run_redshift_query(edited_sql, redshift_engine)
                if results_df is not None:
                    st.dataframe(results_df)