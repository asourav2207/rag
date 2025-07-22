import sys
import os
# Ensure parent dir is in sys.path for package imports
parent_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)
sys.path.append(os.path.abspath(os.path.dirname(__file__)))

os.environ["OMP_NUM_THREADS"] = "1"

import streamlit as st
import pandas as pd
from datetime import datetime
from sentence_transformers import SentenceTransformer
import numpy as np
import faiss
import requests
import json
from sqlalchemy import create_engine, Column, Integer, String, Text, DateTime, text
from sqlalchemy.orm import declarative_base, sessionmaker
import base64
from io import BytesIO
from openpyxl import load_workbook
import concurrent.futures
import re
from urllib.parse import quote_plus
import socket
import asyncio
import hashlib
from functools import lru_cache
import time
import threading
from typing import List, Dict, Tuple, Optional

# ---- App Configuration ----
st.set_page_config(
    page_title="🚀 Redshift AI Assistant",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={
        'Get Help': 'https://github.com/asourav2207/rag',
        'Report a bug': 'https://github.com/asourav2207/rag/issues',
        'About': "# Redshift AI Assistant\nPowered by Multi-Agent AI System"
    }
)

# Custom CSS for modern UI
st.markdown("""
<style>
 /* Import Google Fonts */
 @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

 /* Global Styles */
 .main .block-container {
     padding-top: 2rem;
     padding-bottom: 2rem;
     max-width: 1200px;
 }

 /* Custom Font */
 html, body, [class*="css"] {
     font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
 }

 /* Header Styling */
 .main-header {
     background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
     padding: 2rem;
     border-radius: 20px;
     margin-bottom: 2rem;
     text-align: center;
     box-shadow: 0 10px 30px rgba(0,0,0,0.1);
 }

 .main-header h1 {
     color: white;
     font-size: 2.5rem;
     font-weight: 700;
     margin: 0;
     text-shadow: 0 2px 4px rgba(0,0,0,0.3);
 }

 .main-header p {
     color: rgba(255,255,255,0.9);
     font-size: 1.1rem;
     margin: 0.5rem 0 0 0;
     font-weight: 300;
 }

 /* Mode Toggle Styling */
 .mode-toggle {
     background: white;
     padding: 1.5rem;
     border-radius: 15px;
     margin-bottom: 1.5rem;
     box-shadow: 0 4px 20px rgba(0,0,0,0.08);
     border: 1px solid #e1e5e9;
 }

 /* Agent Status Cards */
 .agent-status-card {
     background: linear-gradient(135deg, #f8fafc 0%, #e2e8f0 100%);
     padding: 1rem;
     border-radius: 12px;
     margin: 0.5rem 0;
     border-left: 4px solid #10b981;
     box-shadow: 0 2px 8px rgba(0,0,0,0.05);
 }

 .agent-status-card.inactive {
     border-left-color: #ef4444;
     background: linear-gradient(135deg, #fef2f2 0%, #fee2e2 100%);
 }

 .agent-status-card.partial {
     border-left-color: #f59e0b;
     background: linear-gradient(135deg, #fffbeb 0%, #fef3c7 100%);
 }

 /* Chat Interface */
 .chat-container {
     background: white;
     border-radius: 20px;
     padding: 1.5rem;
     margin-bottom: 2rem;
     box-shadow: 0 4px 20px rgba(0,0,0,0.08);
     border: 1px solid #e1e5e9;
 }

 /* Input Styling */
 .stTextInput > div > div > input {
     border-radius: 12px;
     border: 2px solid #e1e5e9;
     padding: 0.75rem 1rem;
     font-size: 1rem;
     transition: all 0.3s ease;
 }

 .stTextInput > div > div > input:focus {
     border-color: #667eea;
     box-shadow: 0 0 0 3px rgba(102, 126, 234, 0.1);
 }

 /* Button Styling */
 .stButton > button {
     background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
     color: white;
     border: none;
     border-radius: 12px;
     padding: 0.75rem 2rem;
     font-weight: 600;
     font-size: 1rem;
     transition: all 0.3s ease;
     box-shadow: 0 4px 15px rgba(102, 126, 234, 0.4);
 }

 .stButton > button:hover {
     transform: translateY(-2px);
     box-shadow: 0 8px 25px rgba(102, 126, 234, 0.6);
 }

 /* Secondary Button */
 .secondary-button {
     background: linear-gradient(135deg, #f8fafc 0%, #e2e8f0 100%);
     color: #475569;
     border: 2px solid #e2e8f0;
 }

 .secondary-button:hover {
     background: linear-gradient(135deg, #e2e8f0 0%, #cbd5e1 100%);
     border-color: #cbd5e1;
 }

 /* Success Button */
 .success-button {
     background: linear-gradient(135deg, #10b981 0%, #059669 100%);
     box-shadow: 0 4px 15px rgba(16, 185, 129, 0.4);
 }

 /* Danger Button */
 .danger-button {
     background: linear-gradient(135deg, #ef4444 0%, #dc2626 100%);
     box-shadow: 0 4px 15px rgba(239, 68, 68, 0.4);
 }

 /* Metric Cards */
 .metric-card {
     background: white;
     padding: 1.5rem;
     border-radius: 15px;
     text-align: center;
     box-shadow: 0 4px 20px rgba(0,0,0,0.08);
     border: 1px solid #e1e5e9;
     transition: all 0.3s ease;
 }

 .metric-card:hover {
     transform: translateY(-4px);
     box-shadow: 0 8px 30px rgba(0,0,0,0.12);
 }

 .metric-value {
     font-size: 2rem;
     font-weight: 700;
     color: #667eea;
     margin: 0;
 }

 .metric-label {
     font-size: 0.9rem;
     color: #64748b;
     margin: 0.5rem 0 0 0;
     font-weight: 500;
 }

 /* Sidebar Styling */
 .css-1d391kg {
     background: linear-gradient(180deg, #f8fafc 0%, #ffffff 100%);
 }

 /* Expander Styling */
 .streamlit-expanderHeader {
     background: linear-gradient(135deg, #f8fafc 0%, #e2e8f0 100%);
     border-radius: 12px;
     border: 1px solid #e1e5e9;
     font-weight: 600;
 }

 /* Code Block Styling */
 .stCode {
     border-radius: 12px;
     border: 1px solid #e1e5e9;
 }

 /* Alert Styling */
 .stAlert {
     border-radius: 12px;
     border: none;
     box-shadow: 0 4px 15px rgba(0,0,0,0.08);
 }

 /* Chat Message Styling */
 .stChatMessage {
     border-radius: 15px;
     margin: 1rem 0;
     box-shadow: 0 2px 10px rgba(0,0,0,0.05);
 }

 /* Progress Bar */
 .stProgress .st-bo {
     background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
     border-radius: 10px;
 }

 /* Selectbox Styling */
 .stSelectbox > div > div > div {
     border-radius: 12px;
     border: 2px solid #e1e5e9;
 }

 /* File Uploader */
 .stFileUploader {
     border-radius: 15px;
     border: 2px dashed #e1e5e9;
     padding: 2rem;
     text-align: center;
     transition: all 0.3s ease;
 }

 .stFileUploader:hover {
     border-color: #667eea;
     background: rgba(102, 126, 234, 0.02);
 }

 /* Suggestion Pills */
 .suggestion-pill {
     display: inline-block;
     background: linear-gradient(135deg, #f1f5f9 0%, #e2e8f0 100%);
     color: #475569;
     padding: 0.5rem 1rem;
     margin: 0.25rem;
     border-radius: 25px;
     border: 1px solid #e2e8f0;
     font-size: 0.9rem;
     cursor: pointer;
     transition: all 0.3s ease;
 }

 .suggestion-pill:hover {
     background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
     color: white;
     transform: translateY(-1px);
     box-shadow: 0 4px 12px rgba(102, 126, 234, 0.3);
 }

 /* Status Indicators */
 .status-indicator {
     display: inline-block;
     width: 8px;
     height: 8px;
     border-radius: 50%;
     margin-right: 0.5rem;
 }

 .status-active { background: #10b981; }
 .status-inactive { background: #ef4444; }
 .status-partial { background: #f59e0b; }

 /* Loading Animation */
 @keyframes pulse {
     0% { opacity: 1; }
     50% { opacity: 0.5; }
     100% { opacity: 1; }
 }

 .loading-pulse {
     animation: pulse 1.5s ease-in-out infinite;
 }

 /* Responsive Design */
 @media (max-width: 768px) {
     .main-header h1 {
         font-size: 2rem;
     }

     .main .block-container {
         padding-left: 1rem;
         padding-right: 1rem;
     }
 }
</style>
""", unsafe_allow_html=True)

# Modern Header
st.markdown("""
<div class="main-header">
 <h1>🚀 Redshift AI Assistant</h1>
 <p>Intelligent Multi-Agent System for Database Documentation & Query Generation</p>
</div>
""", unsafe_allow_html=True)

# ---- GitHub Integration Config ----
GITHUB_TOKEN = st.secrets.get("github_token", os.environ.get("GITHUB_TOKEN", ""))
GITHUB_REPO = st.secrets.get("github_repo", "asourav2207/rag")
GITHUB_BRANCH = st.secrets.get("github_branch", "main")
GITHUB_EXCEL_PATH = st.secrets.get("github_excel_path", "excel_files")

# ---- Redshift & Database Setup ----
Base = declarative_base()
DB_PATH = "sqlite:///chat_history.db"

# ---- SQLAlchemy ORM Models ----
class ChatSession(Base):
    __tablename__ = 'chat_session'
    id = Column(Integer, primary_key=True)
    name = Column(String, default="Untitled Chat")
    mode = Column(String, default="Documentation")
    created_at = Column(DateTime, default=datetime.now)

class ChatHistory(Base):
    __tablename__ = 'chat_history'
    id = Column(Integer, primary_key=True)
    question = Column(Text)
    answer = Column(Text)
    context = Column(Text)  # This will now store JSON string for rich context display
    sql_query = Column(Text)
    timestamp = Column(DateTime, default=datetime.now)
    feedback = Column(String)
    session_id = Column(Integer)

engine = create_engine(DB_PATH)
Base.metadata.create_all(engine)
SessionLocal = sessionmaker(bind=engine)

# ---- Redshift Functions ----
@st.cache_resource(show_spinner="Connecting to Redshift...")
def get_redshift_engine():
    try:
        creds = st.secrets["redshift"]
        encoded_password = quote_plus(creds['password'])
        conn_string = (
            f"postgresql+psycopg2://{creds['user']}:{encoded_password}@"
            f"{creds['host']}:{creds['port']}/{creds['database']}"
        )
        engine = create_engine(conn_string, connect_args={'connect_timeout': 10})
        with engine.connect() as conn:
            st.sidebar.success("Redshift connection successful!")
        return engine
    except Exception as e:
        st.sidebar.error(f"Redshift connection failed: {e}")
        return None

@st.cache_data(ttl=3600, show_spinner="Fetching Redshift schema...")
def get_redshift_schema(_engine):
    if _engine is None:
        return "Redshift connection not available."
    schema_query = """
    SELECT table_schema, table_name, column_name, data_type
    FROM information_schema.columns
    WHERE table_schema NOT IN ('information_schema', 'pg_catalog')
    ORDER BY table_schema, table_name, ordinal_position;
    """
    try:
        with _engine.connect() as conn:
            result = conn.execute(text(schema_query))
            schema_df = pd.DataFrame(result.fetchall(), columns=result.keys())
            schema_str = ""
            for (schema, table), group in schema_df.groupby(['table_schema', 'table_name']):
                schema_str += f"Table: {schema}.{table}\n"
                for _, row in group.iterrows():
                    schema_str += f" - {row['column_name']} ({row['data_type']})\n"
                schema_str += "\n"
            return schema_str
    except Exception as e:
        st.error(f"Failed to fetch Redshift schema: {e}")
        return None

def run_redshift_query(sql, engine):
    if engine is None:
        st.error("Cannot run query, Redshift connection is not available.")
        return None
    try:
        with engine.connect() as conn:
            result = conn.execute(text(sql))
            if result.returns_rows:
                return pd.DataFrame(result.fetchall(), columns=result.keys())
            else:
                st.success("Query executed successfully.")
                return pd.DataFrame([{"status": "Success", "rows_affected": result.rowcount}])
    except Exception as e:
        st.error(f"Redshift query error: {e}")
        return None

def redshift_connection_diagnostics():
    """Runs a series of checks to diagnose Redshift connection issues."""
    st.sidebar.subheader("⚙️ Connection Diagnostics")
    try:
        creds = st.secrets["redshift"]
        host = creds.get("host")
        port = int(creds.get("port", 5439))

        if not host:
            st.sidebar.error("`host` not found in `secrets.toml` under `[redshift]`.")
            return

        st.sidebar.markdown(f"**Target Host:** `{host}`")
        st.sidebar.markdown(f"**Target Port:** `{port}`")

        # 1. Check Public IP
        public_ip = "unknown"
        try:
            public_ip_req = requests.get("https://checkip.amazonaws.com", timeout=5)
            public_ip_req.raise_for_status()
            public_ip = public_ip_req.text.strip()
            st.sidebar.markdown(f"**Your Public IP:** `{public_ip}`")
            st.sidebar.info(f"Your Redshift Security Group must allow inbound traffic from this IP on port {port}.")
        except Exception as e:
            st.sidebar.warning(f"Could not determine your public IP: {e}")

        # 2. Try a direct socket connection
        st.sidebar.markdown("---")
        st.sidebar.write("Attempting a direct network connection...")
        try:
            with st.spinner("Connecting socket..."):
                s = socket.create_connection((host, port), timeout=10)
                s.close()
            st.sidebar.success("✅ **Success:** Network path to Redshift is open.")
        except Exception as e:
            st.sidebar.error(f"❌ **Failure:** Network connection timed out.")
            st.sidebar.code(f"Error: {e}", language='text')
            st.sidebar.warning(
                "This confirms a network block. Please check the following in your AWS Console, in this order:"
            )
            st.sidebar.markdown("""
1.  **Route Table**: The subnet for your Redshift workgroup **must** have a route `0.0.0.0/0` pointing to an **Internet Gateway (`igw-xxxx`)**.
2.  **Network ACL (NACL)**: Ensure both Inbound and Outbound rules on the subnet's NACL allow traffic from/to your IP on the correct ports.
3.  **Security Group**: The Redshift workgroup's security group must have an Inbound rule allowing TCP on port 5439 from your current IP.
""")
    except KeyError:
        st.sidebar.error("`[redshift]` section not found in `secrets.toml`.")

def extract_sql_from_answer(answer):
    match = re.search(r'```sql\s*(.*?)\s*```', answer, re.DOTALL | re.IGNORECASE)
    return match.group(1).strip() if match else None

# ---- GitHub, Excel, & Changelog Functions ----
def github_headers():
    return {"Authorization": f"token {GITHUB_TOKEN}", "Accept": "application/vnd.github.v3+json"}

def list_github_excels():
    url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{GITHUB_EXCEL_PATH}?ref={GITHUB_BRANCH}"
    try:
        r = requests.get(url, headers=github_headers(), timeout=10)
        r.raise_for_status()
        return [f for f in r.json() if f["name"].endswith((".xlsx", ".xls"))]
    except Exception:
        return []

def download_excel_from_github(file_info):
    try:
        r = requests.get(file_info["download_url"], timeout=10)
        r.raise_for_status()
        # FIX: Wrap r.content in BytesIO for pandas compatibility
        return pd.read_excel(BytesIO(r.content), engine="openpyxl")
    except Exception as e:
        st.error(f"Error downloading {file_info['name']}: {e}")
        return None

def generate_changelog(existing_df, new_df):
    diffs = []
    for i in range(max(len(existing_df), len(new_df))):
        if i >= len(existing_df):
            diffs.append(f"Row {i+1} added: {new_df.iloc[i].to_dict()}")
            continue
        if i >= len(new_df):
            diffs.append(f"Row {i+1} removed: {existing_df.iloc[i].to_dict()}")
            continue

        old_row, new_row = existing_df.iloc[i], new_df.iloc[i]
        if not old_row.equals(new_row):
            for col in new_df.columns:
                if col in old_row and str(old_row[col]) != str(new_row[col]):
                    diffs.append(f"Row {i+1}, Column '{col}': changed from '{old_row[col]}' to '{new_row[col]}'")
    return diffs

def inject_changelog_to_excel(uploaded_file, changelog_lines, user_id):
    uploaded_file.seek(0)
    wb = load_workbook(uploaded_file)
    ws = wb.create_sheet("Changelog") if "Changelog" not in wb.sheetnames else wb["Changelog"]
    if ws.max_row == 1:
        if ws.cell(row=1, column=1).value not in ["Timestamp", "user", "User", "Change Description"]:
            ws.append(["Timestamp", "User", "Change Description"])
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for line in changelog_lines:
        ws.append([timestamp, user_id, line])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def upload_excel_to_github(file, filename):
    url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{GITHUB_EXCEL_PATH}/{filename}"
    try:
        get_resp = requests.get(url, headers=github_headers(), timeout=10)
        sha = get_resp.json().get("sha") if get_resp.status_code == 200 else None
        file.seek(0)
        content = base64.b64encode(file.read()).decode()
        data = {"message": f"Update {filename}", "content": content, "branch": GITHUB_BRANCH}
        if sha:
            data["sha"] = sha
        put_resp = requests.put(url, headers=github_headers(), data=json.dumps(data), timeout=15)
        put_resp.raise_for_status()
        return True
    except Exception as e:
        st.error(f"GitHub upload failed for {filename}: {e}")
        return False

# ---- RAG (Documentation Mode) Functions ----
@st.cache_resource(show_spinner="Loading documentation model and data...")
def load_and_prepare_rag_data():
    """Enhanced RAG data preparation with better document structuring and dual-index approach"""
    row_docs, field_docs, row_metadata, field_metadata = [], [], [], []
    excel_files = list_github_excels()
    if not excel_files:
        st.warning("No Excel documentation files found in GitHub.")
        return None, None, [], [], None, [], []

    for file_info in excel_files:
        df = download_excel_from_github(file_info)
        if df is None:
            continue
        table_name = os.path.splitext(file_info["name"])[0]

        try:
            workbook = load_workbook(BytesIO(requests.get(file_info["download_url"], timeout=10).content))
            if "Changelog" in workbook.sheetnames:
                st.info(f"Skipping 'Changelog' sheet in {file_info['name']} for RAG processing.")
                df = pd.read_excel(BytesIO(requests.get(file_info["download_url"], timeout=10).content),
                                   sheet_name=workbook.sheetnames[0], engine="openpyxl")
            if "Changelog" in df.columns:
                df = df.drop(columns=["Changelog"], errors='ignore')
        except Exception as e:
            st.warning(f"Could not check sheets for {file_info['name']}: {e}. Proceeding with default sheet.")

        for idx, row in df.iterrows():
            non_empty_fields = []
            field_summaries = []

            for col in df.columns:
                value = str(row.get(col, '')).strip()
                if value and value.lower() not in ['nan', 'none', '']:
                    field_doc = f"{col}: {value}"
                    field_docs.append(field_doc)
                    field_metadata.append({
                        'table': table_name,
                        'column': col,
                        'value': value,
                        'row_index': idx,
                        'field_name': col.lower().replace(' ', '_')
                    })
                    non_empty_fields.append(f"{col}: {value}")
                    if any(keyword in col.lower() for keyword in ['name', 'description', 'purpose', 'definition']):
                        field_summaries.append(f"{value}")

            if non_empty_fields:
                if field_summaries:
                    row_doc = f"{' | '.join(field_summaries[:2])}\n{chr(10).join(non_empty_fields)}\nSource: {table_name} (Row {idx})"
                else:
                    row_doc = f"{chr(10).join(non_empty_fields)}\nSource: {table_name} (Row {idx})"

                row_docs.append(row_doc)
                row_metadata.append({
                    'table': table_name,
                    'row_index': idx,
                    'content': chr(10).join(non_empty_fields),
                    'original_row_data': row.to_dict(),
                    'field_count': len(non_empty_fields),
                    'has_description': any('description' in col.lower() for col in df.columns if str(row.get(col, '')).strip())
                })

    if not row_docs:
        return None, None, [], [], None, [], []

    model = SentenceTransformer('all-MiniLM-L6-v2')
    with st.spinner("Generating optimized embeddings..."):
        row_embeddings = model.encode(row_docs, show_progress_bar=False, normalize_embeddings=True)
        field_embeddings = model.encode(field_docs, show_progress_bar=False, normalize_embeddings=True) if field_docs else np.array([])
        dim = model.get_sentence_embedding_dimension()

        if len(row_docs) > 100:
            quantizer = faiss.IndexFlatIP(dim)
            row_index = faiss.IndexIVFFlat(quantizer, dim, min(int(np.sqrt(len(row_docs))), 100))
            row_index.train(row_embeddings.astype('float32'))
            row_index.add(row_embeddings.astype('float32'))
        else:
            row_index = faiss.IndexFlatIP(dim)
            row_index.add(row_embeddings.astype('float32'))

        if len(field_docs) > 0:
            field_index = faiss.IndexFlatIP(dim)
            field_index.add(field_embeddings.astype('float32'))
        else:
            field_index = None

    st.success(f"✅ Enhanced indexing complete: {len(row_docs)} row documents, {len(field_docs)} field documents")
    return model, row_index, row_docs, row_metadata, field_index, field_docs, field_metadata

def preprocess_query(query: str) -> Tuple[str, List[str]]:
    """Enhanced query preprocessing with keyword extraction and expansion"""
    normalized_query = query.lower().strip()
    key_terms = []
    field_patterns = [
        r'\b(\w+)\s+field\b', r'\b(\w+)\s+column\b', r'\bfield\s+(\w+)\b',
        r'\bcolumn\s+(\w+)\b', r'\b(\w+_\w+)\b'
    ]
    for pattern in field_patterns:
        matches = re.findall(pattern, normalized_query)
        key_terms.extend(matches)

    quoted_terms = re.findall(r'"([^"]+)"', query)
    key_terms.extend(quoted_terms)
    expansions = {
        'desc': 'description', 'def': 'definition', 'spec': 'specification',
        'amt': 'amount', 'qty': 'quantity', 'num': 'number', 'id': 'identifier',
        'cd': 'code', 'dt': 'date'
    }
    expanded_query = normalized_query
    for abbr, full in expansions.items():
        expanded_query = re.sub(rf'\b{abbr}\b', f"{abbr} {full}", expanded_query)
    return expanded_query, key_terms

def enhanced_retrieve_relevant_docs(query: str, model, row_index, row_docs, row_metadata,
                                    field_index=None, field_docs=None, field_metadata=None, k=10):
    """Enhanced document retrieval with hybrid search and intelligent ranking"""
    if not row_docs:
        return []
    processed_query, key_terms = preprocess_query(query)
    query_hash = hashlib.md5(f"{processed_query}_{k}".encode()).hexdigest()
    cache_key = f"enhanced_docs_{query_hash}"
    cached_results = response_cache.get(processed_query, cache_key)
    if cached_results:
        return cached_results['matches']

    all_candidates = []
    query_emb = model.encode([processed_query], normalize_embeddings=True)
    search_k = min(k * 3, len(row_docs))
    scores, indices = row_index.search(query_emb.astype('float32'), search_k)
    for i, (score, idx) in enumerate(zip(scores[0], indices[0])):
        if idx >= 0 and score > 0.1:
            all_candidates.append({
                'doc': row_docs[idx], 'metadata': row_metadata[idx],
                'score': float(score), 'rank': i, 'source': 'semantic_row'
            })

    if field_index is not None and field_docs and len(key_terms) > 0:
        field_scores, field_indices = field_index.search(query_emb.astype('float32'), min(k * 2, len(field_docs)))
        for score, idx in zip(field_scores[0], field_indices[0]):
            if idx >= 0 and score > 0.15:
                field_meta = field_metadata[idx]
                row_idx = next((i for i, meta in enumerate(row_metadata)
                                if meta['table'] == field_meta['table'] and
                                meta['row_index'] == field_meta['row_index']), None)
                if row_idx is not None:
                    all_candidates.append({
                        'doc': row_docs[row_idx], 'metadata': row_metadata[row_idx],
                        'score': float(score * 1.1), 'rank': 0,
                        'source': 'semantic_field', 'matching_field': field_meta['column']
                    })

    for term in key_terms:
        if len(term) > 2:
            for i, (doc, meta) in enumerate(zip(row_docs, row_metadata)):
                if term.lower() in doc.lower():
                    term_count = doc.lower().count(term.lower())
                    keyword_score = min(term_count * 0.3, 1.0)
                    all_candidates.append({
                        'doc': doc, 'metadata': meta, 'score': keyword_score,
                        'rank': 0, 'source': 'keyword', 'matching_term': term
                    })

    doc_scores = {}
    for candidate in all_candidates:
        doc_key = f"{candidate['metadata']['table']}_{candidate['metadata']['row_index']}"
        if doc_key not in doc_scores:
            doc_scores[doc_key] = {
                'doc': candidate['doc'], 'metadata': candidate['metadata'],
                'total_score': 0, 'source_scores': {}, 'matching_info': []
            }
        weight = {'semantic_row': 1.0, 'semantic_field': 1.3, 'keyword': 1.5}.get(candidate['source'], 1.0)
        doc_scores[doc_key]['total_score'] += candidate['score'] * weight
        doc_scores[doc_key]['source_scores'][candidate['source']] = candidate['score']
        if 'matching_field' in candidate:
            doc_scores[doc_key]['matching_info'].append(f"Field: {candidate['matching_field']}")
        if 'matching_term' in candidate:
            doc_scores[doc_key]['matching_info'].append(f"Term: {candidate['matching_term']}")

    ranked_results = []
    for doc_key, doc_info in doc_scores.items():
        boost = 1.0
        if doc_info['metadata'].get('has_description', False):
            boost += 0.1
        field_count = doc_info['metadata'].get('field_count', 0)
        if field_count > 5:
            boost += 0.05
        if 'keyword' in doc_info['source_scores']:
            boost += 0.2
        final_score = doc_info['total_score'] * boost
        ranked_results.append({
            'doc': doc_info['doc'], 'metadata': doc_info['metadata'],
            'score': final_score, 'matching_info': doc_info['matching_info']
        })

    ranked_results.sort(key=lambda x: x['score'], reverse=True)
    final_matches = [(result['doc'], result['metadata']) for result in ranked_results[:k]]
    response_cache.set(processed_query, {'matches': final_matches}, cache_key)
    if len(final_matches) > 0:
        st.sidebar.info(f"🎯 Retrieved {len(final_matches)} documents (Score: {ranked_results[0]['score']:.3f})")
    return final_matches

# ---- LLM & Chat Functions ----
def call_ollama(prompt, timeout=90):
    try:
        response = requests.post("http://127.0.0.1:11434/api/generate",
                                 json={"model": "llama3.2", "prompt": prompt, "stream": False}, timeout=timeout)
        response.raise_for_status()
        return response.json().get("response", "").strip()
    except requests.exceptions.RequestException as e:
        return f"Error contacting Ollama: {e}"

def get_rag_answer(user_query, rag_data, history):
    model, row_index, row_docs, row_metadata, field_index, field_docs, field_metadata = rag_data
    if not model or not row_index or not row_docs or not row_metadata:
        return "Documentation data not loaded or indexed. Please ensure Excel files are uploaded and data is available.", "", None

    ambiguous_terms = ["this field", "it", "the field", "that field"]
    if any(term in user_query.lower() for term in ambiguous_terms) and st.session_state.get("last_referenced_field"):
        user_query += f" (referring to {st.session_state.last_referenced_field})"

    matches = enhanced_retrieve_relevant_docs(
        user_query, model, row_index, row_docs, row_metadata,
        field_index, field_docs, field_metadata, k=8
    )
    if matches:
        first_match_content = matches[0][1].get('content', '')
        m = re.search(r"([a-zA-Z0-9_]+)\s*:", first_match_content)
        if m:
            st.session_state["last_referenced_field"] = m.group(1)

    context_history = history_manager.get_context_window(history, window_size=3)
    context_for_llm = "\n---\n".join([doc for doc, meta in matches][:5])
    conversation_history = "".join([f"User: {h.question}\nAssistant: {h.answer}\n" for h in context_history if h.answer != "Thinking..."])
    prompt = f"""You are a helpful and precise documentation assistant. Your primary goal is to answer the user's question STRICTLY based on the provided "Documentation Context".

IMPORTANT: The documentation context is ranked by relevance - the FIRST context entry is the MOST RELEVANT to the user's question. Pay special attention to the first few context entries.

If the answer is not directly available or cannot be reasonably inferred from the context, you MUST state: "I cannot answer this question based on the provided documentation."
Do NOT use any outside knowledge or make assumptions. Be concise and direct.

When you find relevant information in the context, reference which context entry it came from (e.g., "Based on the first context entry..." or "According to the documentation...").

Conversation History:
{conversation_history}

Documentation Context (ranked by relevance):
{context_for_llm}

User Query: {user_query}

Answer:"""
    answer = async_llm_call(prompt)
    return answer, json.dumps([m[1] for m in matches]), None

def get_sql_answer(user_query, schema_str, history):
    context_history = history_manager.get_context_window(history, window_size=3)
    conversation_history = "".join([f"User: {h.question}\nAssistant:\n{h.answer}\n" for h in context_history if h.answer != "Thinking..."])
    prompt = f"""You are an expert Redshift SQL analyst. Convert the user's question into a correct Redshift SQL query using the provided schema and history.
Conversation History:
{conversation_history}
Redshift Schema:
{schema_str}
User Question: {user_query}
Generate ONLY the SQL query in a single SQL code block. Example: ```sql\nSELECT * FROM my_table;\n```"""
    response = async_llm_call(prompt)
    sql_query = extract_sql_from_answer(response)
    answer = "I have translated your request into the SQL query below. You can review, edit, and run it."
    if not sql_query:
        answer = "I was unable to generate a SQL query. Please try rephrasing."
    return answer, schema_str, sql_query

def generate_followup_suggestions_async(user_prompt, system_response, context):
    def _llm_suggest():
        context_summary = ""
        try:
            if context and context.strip().startswith('['):
                parsed_context = json.loads(context)
                if parsed_context:
                    context_summary = "Relevant Tables/Fields from Documentation:\n" + "\n".join(
                        [f"- {ctx.get('table', '')} (Row {ctx.get('row_index', '')})" for ctx in parsed_context[:2]]
                    )
            else:
                context_summary = f"Context snippet: {str(context)[:200]}..."
        except (json.JSONDecodeError, TypeError):
            context_summary = f"Context snippet: {str(context)[:200]}..."

        prompt = f"""Given the user question, system answer, and relevant context, generate 2-3 concise, direct follow-up questions a user might ask next. Phrase each as a question.
User question: {user_prompt}
System answer: {system_response}
{context_summary}
List the follow-up questions as bullet points."""
        suggestions = call_ollama(prompt, timeout=20)
        if suggestions.startswith("Error contacting Ollama"):
            return ["(Suggestions: Ollama not reachable)"]
        return [line.lstrip('-• ').strip() for line in suggestions.split('\n') if line.strip()]

    with concurrent.futures.ThreadPoolExecutor() as executor:
        future = executor.submit(_llm_suggest)
        try:
            return future.result(timeout=25)
        except concurrent.futures.TimeoutError:
            return ["(Suggestions timed out)"]
        except Exception as e:
            return [f"(Error generating suggestions: {e})"]

def save_chat(question, answer, context, sql, session_id):
    db = SessionLocal()
    entry = ChatHistory(question=question, answer=answer, context=context, sql_query=sql, session_id=session_id)
    db.add(entry)
    db.commit()
    db.close()

def get_chat_history(session_id):
    db = SessionLocal()
    entries = db.query(ChatHistory).filter_by(session_id=session_id).order_by(ChatHistory.timestamp).all()
    db.close()
    return entries

# Initialize session state variables
if 'selected_session_id' not in st.session_state:
    st.session_state.selected_session_id = None
if 'history' not in st.session_state:
    st.session_state.history = []
if 'user_input' not in st.session_state:
    st.session_state.user_input = ""
if 'rerun_triggered_by_suggestion' not in st.session_state:
    st.session_state.rerun_triggered_by_suggestion = False
if 'suggestion_counter' not in st.session_state:
    st.session_state.suggestion_counter = 0

# ---- Advanced Caching & Memory Management ----
class ResponseCache:
    """Thread-safe response cache with TTL and LRU eviction"""
    def __init__(self, max_size=100, ttl_seconds=3600):
        self.cache = {}
        self.access_times = {}
        self.max_size = max_size
        self.ttl_seconds = ttl_seconds
        self.lock = threading.Lock()

    def _generate_key(self, query: str, context_hash: str = "") -> str:
        """Generate a hash key for caching"""
        combined = f"{query}{context_hash}"
        return hashlib.md5(combined.encode()).hexdigest()

    def get(self, query: str, context_hash: str = "") -> Optional[Dict]:
        """Get cached response if valid"""
        key = self._generate_key(query, context_hash)
        with self.lock:
            if key in self.cache:
                cached_item = self.cache[key]
                if time.time() - cached_item['timestamp'] < self.ttl_seconds:
                    self.access_times[key] = time.time()
                    return cached_item['data']
                else:
                    del self.cache[key]
                    del self.access_times[key]
            return None

    def set(self, query: str, data: Dict, context_hash: str = ""):
        """Cache response with LRU eviction"""
        key = self._generate_key(query, context_hash)
        with self.lock:
            if len(self.cache) >= self.max_size:
                oldest_key = min(self.access_times.keys(), key=lambda k: self.access_times[k])
                del self.cache[oldest_key]
                del self.access_times[oldest_key]
            self.cache[key] = {'data': data, 'timestamp': time.time()}
            self.access_times[key] = time.time()

@st.cache_resource
def get_response_cache():
    return ResponseCache()

response_cache = get_response_cache()

# ---- Async Processing Functions ----
async def async_embedding_generation(texts: List[str], model) -> np.ndarray:
    """Asynchronously generate embeddings in batches"""
    def encode_batch(batch):
        return model.encode(batch, show_progress_bar=False)

    batch_size = 50
    batches = [texts[i:i + batch_size] for i in range(0, len(texts), batch_size)]
    with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
        futures = [executor.submit(encode_batch, batch) for batch in batches]
        results = [future.result() for future in concurrent.futures.as_completed(futures)]
    return np.vstack(results) if results else np.array([])

def async_llm_call(prompt: str, timeout: int = 90) -> str:
    """Async wrapper for LLM calls with caching"""
    cache_key = hashlib.md5(prompt.encode()).hexdigest()[:16]
    cached_response = response_cache.get(prompt, cache_key)
    if cached_response:
        return cached_response['response']

    def _call_llm():
        try:
            response = requests.post("http://127.0.0.1:11434/api/generate",
                                     json={"model": "llama3.2", "prompt": prompt, "stream": False},
                                     timeout=timeout)
            response.raise_for_status()
            return response.json().get("response", "").strip()
        except requests.exceptions.RequestException as e:
            return f"Error contacting Ollama: {e}"

    with concurrent.futures.ThreadPoolExecutor() as executor:
        future = executor.submit(_call_llm)
        try:
            result = future.result(timeout=timeout)
            response_cache.set(prompt, {'response': result}, cache_key)
            return result
        except concurrent.futures.TimeoutError:
            return f"LLM call timed out after {timeout} seconds"

@lru_cache(maxsize=128)
def cached_document_retrieval(query_hash: str, k: int = 5) -> str:
    """Cached document retrieval results"""
    return f"cached_retrieval_{query_hash}_{k}"

def efficient_retrieve_relevant_docs(query: str, model, row_index, row_docs, row_metadata, k=5):
    """Enhanced document retrieval with caching"""
    if not row_docs:
        return []
    query_hash = hashlib.md5(query.encode()).hexdigest()
    cache_key = f"docs_{query_hash}_{k}"
    cached_results = response_cache.get(query, cache_key)
    if cached_results:
        return cached_results['matches']

    query_emb = model.encode([query])
    distances, indices = row_index.search(query_emb, k)
    all_matches = []
    for i, dist in zip(indices[0], distances[0]):
        if dist < 1.5:
            all_matches.append((row_docs[i], row_metadata[i], dist))

    all_matches.sort(key=lambda x: x[2])
    unique_matches = [(doc, meta) for doc, meta, dist in all_matches[:k]]
    response_cache.set(query, {'matches': unique_matches}, cache_key)
    return unique_matches

class ChatHistoryManager:
    """Efficient chat history management with windowing"""
    @staticmethod
    def get_context_window(history: List, window_size: int = 5) -> List:
        """Get recent context window for better memory efficiency"""
        if not history:
            return []
        filtered_history = [h for h in history if h.answer != "Thinking..."]
        return filtered_history[-window_size:] if len(filtered_history) > window_size else filtered_history

    @staticmethod
    def compress_old_history(session_id: int, keep_recent: int = 50):
        """Compress old chat history to save memory"""
        db = SessionLocal()
        try:
            total_entries = db.query(ChatHistory).filter_by(session_id=session_id).count()
            if total_entries > keep_recent * 2:
                recent_entries = db.query(ChatHistory).filter_by(session_id=session_id)\
                    .order_by(ChatHistory.timestamp.desc()).limit(keep_recent).all()
                recent_ids = [e.id for e in recent_entries]
                db.query(ChatHistory).filter_by(session_id=session_id)\
                    .filter(~ChatHistory.id.in_(recent_ids)).delete(synchronize_session=False)
                db.commit()
        finally:
            db.close()

history_manager = ChatHistoryManager()

# ---- Multi-Agent System Integration ----
try:
    from agents.multi_agent_integration import create_multi_agent_system, process_query_with_agents
    MULTI_AGENT_AVAILABLE = True
except ImportError:
    MULTI_AGENT_AVAILABLE = False
    st.warning("Multi-agent system not available. Using fallback mode.")

# ---- Enhanced Redshift Integration ----
try:
    from agents.metadata_manager import SchemaMetadataManager
    from agents.redshift_agent import EnhancedRedshiftQueryAgent
    ENHANCED_REDSHIFT_AVAILABLE = True
except ImportError:
    ENHANCED_REDSHIFT_AVAILABLE = False
    st.warning("Enhanced Redshift features not available.")

@st.cache_resource
def initialize_multi_agent_system(_rag_data, _redshift_engine, _schema_info):
    """Initialize the multi-agent system with caching"""
    if not MULTI_AGENT_AVAILABLE:
        return None
    try:
        return create_multi_agent_system(
            rag_data=_rag_data,
            redshift_engine=_redshift_engine,
            schema_info=_schema_info,
            llm_client=async_llm_call,
            cache_manager=response_cache
        )
    except Exception as e:
        st.error(f"Failed to initialize multi-agent system: {e}")
        return None

@st.cache_resource
def initialize_metadata_manager():
    """Initialize the metadata manager for enhanced Redshift features"""
    if not ENHANCED_REDSHIFT_AVAILABLE:
        return None
    try:
        return SchemaMetadataManager()
    except Exception as e:
        st.error(f"Failed to initialize metadata manager: {e}")
        return None

def display_metadata_overview(metadata_manager):
    """Display metadata overview and statistics in sidebar"""
    if not metadata_manager:
        return
    try:
        import sqlite3
        conn = sqlite3.connect(metadata_manager.metadata_db_path)
        table_count = pd.read_sql_query(
            "SELECT COUNT(DISTINCT table_name) as count FROM schema_metadata", conn
        )['count'].iloc[0] if not pd.read_sql_query("SELECT name FROM sqlite_master WHERE type='table' AND name='schema_metadata'", conn).empty else 0

        if table_count > 0:
            column_count = pd.read_sql_query(
                "SELECT COUNT(*) as count FROM schema_metadata WHERE column_name IS NOT NULL", conn
            )['count'].iloc[0]
            relationship_count = pd.read_sql_query(
                "SELECT COUNT(*) as count FROM table_relationships", conn
            )['count'].iloc[0]
            pattern_count = pd.read_sql_query(
                "SELECT COUNT(*) as count FROM query_patterns", conn
            )['count'].iloc[0]
            st.sidebar.subheader("📊 Enhanced Schema Metadata")
            col1, col2 = st.sidebar.columns(2)
            with col1:
                st.metric("Tables", table_count)
                st.metric("Relationships", relationship_count)
            with col2:
                st.metric("Columns", column_count)
                st.metric("Query Patterns", pattern_count)
            quality_score = min((column_count / max(table_count * 5, 1)) * 100, 100)
            st.sidebar.metric(
                "Metadata Quality",
                f"{quality_score:.0f}%",
                help="Based on description coverage and relationships"
            )
        conn.close()
    except Exception as e:
        st.sidebar.warning(f"Could not load metadata overview: {e}")

def initialize_enhanced_redshift_metadata(redshift_engine, metadata_manager):
    """Initialize enhanced metadata with descriptions and relationships"""
    if not metadata_manager or not redshift_engine:
        return {"error": "Required components not available"}

    def simple_llm_client(prompt):
        try:
            response = requests.post("http://127.0.0.1:11434/api/generate",
                                     json={"model": "llama3.2", "prompt": prompt, "stream": False},
                                     timeout=30)
            response.raise_for_status()
            generated_text = response.json().get("response", "").strip()
            if not generated_text or "error" in generated_text.lower():
                if "table" in prompt.lower():
                    return "Data table containing structured information for business operations."
                elif "column" in prompt.lower():
                    column_name = ""
                    if "Column: " in prompt:
                        column_name = prompt.split("Column: ")[1].split(" ")[0]
                    return f"Database field {column_name} storing specific data values."
                else:
                    return "Database element with business relevance."
            return generated_text
        except Exception:
            if "id" in prompt.lower(): return "Unique identifier field for record identification"
            if "name" in prompt.lower(): return "Name or label field for descriptive purposes"
            if "date" in prompt.lower() or "time" in prompt.lower(): return "Date/time field for temporal data tracking"
            if "amount" in prompt.lower() or "cost" in prompt.lower(): return "Monetary amount or cost field"
            if "status" in prompt.lower(): return "Status field indicating current state or condition"
            return "Database field containing business data"

    try:
        result = metadata_manager.extract_and_store_schema(redshift_engine, simple_llm_client)
        return result
    except Exception as e:
        return {"error": str(e)}

async def process_enhanced_redshift_query(user_query: str, redshift_engine, schema_info, metadata_manager, history):
    """Process Redshift queries using enhanced metadata-aware agent"""
    if not ENHANCED_REDSHIFT_AVAILABLE or not metadata_manager:
        return get_sql_answer(user_query, schema_info, history)
    try:
        from agents.orchestrator_agent import QueryContext, QueryType
        query_context = QueryContext(
            user_query=user_query, query_type=QueryType.UNCLEAR, confidence=0.0,
            extracted_entities={}, conversation_history=history[-5:] if history else [],
            session_metadata={"execute_query": False}, timestamp=datetime.now()
        )

        def simple_llm_client(prompt):
            return async_llm_call(prompt)

        enhanced_agent = EnhancedRedshiftQueryAgent(redshift_engine, schema_info, simple_llm_client)
        response = await enhanced_agent.process_query(query_context)

        if response.success:
            sql_query = response.context_data.get('sql_query', '')
            explanation = response.response_text
            context_data = {
                "query_type": response.context_data.get('query_type', 'unknown'),
                "confidence": response.confidence,
                "metadata_quality": response.context_data.get('metadata_quality', {}),
                "tables_found": response.metadata.get('tables_found', 0),
                "relationships_used": response.metadata.get('relationships_used', 0),
                "enhanced": True
            }
            return explanation, json.dumps(context_data), sql_query
        else:
            return get_sql_answer(user_query, schema_info, history)
    except Exception as e:
        st.warning(f"Enhanced processing failed: {e}. Using fallback mode.")
        return get_sql_answer(user_query, schema_info, history)

def display_enhanced_schema_explorer(metadata_manager, user_query=""):
    """Display enhanced schema information with search capabilities"""
    if not metadata_manager:
        st.info("Enhanced schema features not available. Initialize metadata first.")
        return
    st.subheader("🔍 Enhanced Schema Explorer")
    search_term = st.text_input(
        "Search tables, columns, or descriptions:",
        value=user_query,
        placeholder="e.g., user, customer, payment, claims",
        key="schema_search"
    )
    if search_term:
        try:
            enhanced_schema = metadata_manager.get_enhanced_schema_for_query(
                search_term,
                {"tables": [], "fields": [], "keywords": search_term.split()}
            )
            if enhanced_schema.get('relevant_metadata'):
                tables_data = {}
                for meta in enhanced_schema['relevant_metadata']:
                    table_key = f"{meta['schema_name']}.{meta['table_name']}"
                    if table_key not in tables_data:
                        tables_data[table_key] = {
                            'description': meta.get('business_context', ''), 'columns': []
                        }
                    if meta.get('column_name'):
                        tables_data[table_key]['columns'].append({
                            'name': meta['column_name'], 'type': meta['data_type'],
                            'nullable': meta['is_nullable'], 'description': meta.get('generated_description', '')
                        })
                for table_name, table_info in tables_data.items():
                    with st.expander(f"📋 {table_name}", expanded=True):
                        if table_info['description']:
                            st.info(f"**Description:** {table_info['description']}")
                        if table_info['columns']:
                            df_columns = pd.DataFrame(table_info['columns'])
                            st.dataframe(df_columns, use_container_width=True)
                if enhanced_schema.get('relationships'):
                    st.subheader("🔗 Table Relationships")
                    for rel in enhanced_schema['relationships']:
                        st.write(
                            f"**{rel['source_table']}.{rel['source_column']}** → "
                            f"**{rel['target_table']}.{rel['target_column']}** "
                            f"({rel['relationship_type']}, confidence: {rel['confidence_score']:.1f})"
                        )
                if enhanced_schema.get('suggested_patterns'):
                    st.subheader("💡 Suggested Query Patterns")
                    for pattern in enhanced_schema['suggested_patterns'][:3]:
                        with st.expander(f"Pattern: {pattern['pattern_description']}"):
                            st.code(pattern['sql_template'], language='sql')
            else:
                st.info("No matching tables or columns found. Try a different search term.")
        except Exception as e:
            st.error(f"Error searching schema: {e}")

def display_query_analytics(metadata_manager):
    """Display query analytics and patterns"""
    if not metadata_manager:
        st.info("Enhanced analytics not available. Initialize metadata first.")
        return
    st.subheader("📊 Query Analytics & Patterns")
    try:
        import sqlite3
        conn = sqlite3.connect(metadata_manager.metadata_db_path)
        patterns_df = pd.read_sql_query("""
        SELECT pattern_description, usage_count, last_used
        FROM query_patterns WHERE usage_count > 0
        ORDER BY usage_count DESC LIMIT 10
        """, conn)
        if not patterns_df.empty:
            st.subheader("🏆 Most Used Query Patterns")
            st.dataframe(patterns_df, use_container_width=True)

        relationships_df = pd.read_sql_query("""
        SELECT source_table, source_column, target_table, target_column,
        relationship_type, confidence_score
        FROM table_relationships ORDER BY last_updated DESC LIMIT 10
        """, conn)
        if not relationships_df.empty:
            st.subheader("🔗 Detected Table Relationships")
            st.dataframe(relationships_df, use_container_width=True)
        conn.close()
    except Exception as e:
        st.error(f"Error loading analytics: {e}")

# ---- Main Application Logic ----
redshift_engine = get_redshift_engine()
schema_info = get_redshift_schema(redshift_engine) if redshift_engine else None
rag_data = load_and_prepare_rag_data()
metadata_manager = initialize_metadata_manager()

multi_agent_system = None
if MULTI_AGENT_AVAILABLE and rag_data[0] is not None:
    multi_agent_system = initialize_multi_agent_system(rag_data, redshift_engine, schema_info)

with st.sidebar:
    st.header("🔧 Configuration")
    mode = st.selectbox(
        "Select Mode:",
        ["📖 Documentation Assistant", "🗄️ SQL Query Assistant", "🔍 Enhanced Schema Explorer"],
        key="mode_selector"
    )
    st.subheader("💬 Chat Sessions")
    new_session_name = st.text_input("New Session Name:", placeholder="Enter session name")
    if st.button("➕ Create Session") and new_session_name.strip():
        db = SessionLocal()
        session_mode = "Documentation" if "Documentation" in mode else "SQL"
        new_session = ChatSession(name=new_session_name.strip(), mode=session_mode)
        db.add(new_session)
        db.commit()
        st.session_state.selected_session_id = new_session.id
        db.close()
        st.rerun()

    db = SessionLocal()
    sessions = db.query(ChatSession).order_by(ChatSession.created_at.desc()).all()
    db.close()

    if sessions:
        session_options = {f"{s.name} ({s.mode})": s.id for s in sessions}
        if st.session_state.selected_session_id is None:
            st.session_state.selected_session_id = sessions[0].id
        selected_session_name = st.selectbox(
            "Select Session:",
            list(session_options.keys()),
            index=0 if st.session_state.selected_session_id is None else
            list(session_options.values()).index(st.session_state.selected_session_id)
            if st.session_state.selected_session_id in session_options.values() else 0
        )
        st.session_state.selected_session_id = session_options[selected_session_name]
        if st.session_state.selected_session_id:
            st.session_state.history = get_chat_history(st.session_state.selected_session_id)

    if metadata_manager:
        display_metadata_overview(metadata_manager)

    st.subheader("📁 Documentation Files")
    excel_files = list_github_excels()
    if excel_files:
        st.write(f"📄 Found {len(excel_files)} Excel files in GitHub")
        for file_info in excel_files[:5]:
            st.text(f"• {file_info['name']}")
        if len(excel_files) > 5:
            st.text(f"... and {len(excel_files) - 5} more")
    else:
        st.warning("No Excel files found in GitHub repository")

    uploaded_file = st.file_uploader("Upload Excel Documentation", type=["xlsx"], key="doc_upload")
    if uploaded_file:
        if st.button("📤 Upload to GitHub"):
            user_id = st.text_input("Your User ID:", placeholder="Enter your ID for changelog")
            if user_id.strip():
                existing_files = [f for f in excel_files if f["name"] == uploaded_file.name]
                if existing_files:
                    existing_df = download_excel_from_github(existing_files[0])
                    new_df = pd.read_excel(uploaded_file, engine="openpyxl")
                    changelog = generate_changelog(existing_df, new_df)
                    if changelog:
                        updated_file = inject_changelog_to_excel(uploaded_file, changelog, user_id.strip())
                        if upload_excel_to_github(updated_file, uploaded_file.name):
                            st.success("✅ File uploaded with changelog!")
                            st.cache_resource.clear()
                            st.rerun()
                        else:
                            st.error("❌ Upload failed")
                    else:
                        st.info("No changes detected")
                else:
                    if upload_excel_to_github(uploaded_file, uploaded_file.name):
                        st.success("✅ New file uploaded!")
                        st.cache_resource.clear()
                        st.rerun()
                    else:
                        st.error("❌ Upload failed")

    st.markdown("---")
    st.subheader("⚙️ Troubleshooting")
    if st.button("Redshift Connection Diagnostics"):
        redshift_connection_diagnostics()

if mode == "📖 Documentation Assistant":
    st.subheader("📖 Documentation Assistant")
    st.write("Ask questions about your database documentation and get intelligent answers.")
    if rag_data[0] is None:
        st.warning("⚠️ No documentation data loaded. Please upload Excel files to GitHub.")
        st.stop()

    user_input = st.chat_input("Ask about documentation (e.g., 'What does user_id mean?', 'How do I join claims and members?')")
    if user_input and user_input.strip():
        with st.spinner("🧠 Processing your question..."):
            if multi_agent_system:
                result = asyncio.run(process_query_with_agents(multi_agent_system, user_input.strip(), st.session_state.history))
                answer = result.get("response_text", "")
                context = json.dumps(result.get("context_data", {}))
                sql_query = result.get("sql_query", None)
            else:
                answer, context, sql_query = get_rag_answer(user_input.strip(), rag_data, st.session_state.history)
            save_chat(user_input.strip(), answer, context, sql_query, st.session_state.selected_session_id)
            st.session_state.history = get_chat_history(st.session_state.selected_session_id)
            st.rerun()

    for entry in reversed(st.session_state.history):
        with st.chat_message("user"):
            st.markdown(entry.question)
        with st.chat_message("assistant"):
            st.markdown("#### 💬 Answer:")
            st.success(entry.answer)
            if entry == st.session_state.history[-1]:
                with st.expander("🔍 Follow-up Suggestions", expanded=False):
                    suggestions = generate_followup_suggestions_async(entry.question, entry.answer, entry.context)
                    for i, suggestion in enumerate(suggestions[:3]):
                        if st.button(f"💡 {suggestion}", key=f"suggestion_{entry.id}_{i}"):
                            st.session_state.user_input = suggestion
                            st.rerun()
            if entry.context:
                with st.expander("📄 Retrieved Context", expanded=False):
                    try:
                        context_data = json.loads(entry.context)
                        for i, ctx in enumerate(context_data[:3]):
                            st.markdown(f"**Source {i+1}:** {ctx.get('table', 'Unknown')} (Row {ctx.get('row_index', 'N/A')})")
                            st.code(ctx.get('content', ''), language='text')
                    except (json.JSONDecodeError, TypeError):
                        st.code(entry.context, language='text')

elif mode == "🗄️ SQL Query Assistant":
    st.subheader("🗄️ SQL Query Assistant")
    st.write("Generate Redshift SQL queries from natural language descriptions.")
    if not redshift_engine:
        st.error("❌ Redshift connection not available. Please check your configuration.")
        st.stop()

    use_enhanced = st.checkbox(
        "🚀 Use Enhanced Query Processing",
        value=ENHANCED_REDSHIFT_AVAILABLE,
        disabled=not ENHANCED_REDSHIFT_AVAILABLE,
        help="Uses advanced metadata and relationship awareness for better query generation"
    )
    if use_enhanced and metadata_manager:
        if st.button("🔄 Initialize Enhanced Metadata"):
            with st.spinner("Analyzing schema and generating metadata..."):
                result = initialize_enhanced_redshift_metadata(redshift_engine, metadata_manager)
                if "error" not in result:
                    st.success("✅ Enhanced metadata initialized successfully!")
                else:
                    st.error(f"❌ Metadata initialization failed: {result['error']}")

    user_input = st.chat_input("Describe the SQL query you need (e.g., 'Show me all active users', 'Join claims with member data')")
    if user_input and user_input.strip():
        with st.spinner("🧠 Generating SQL query..."):
            if multi_agent_system:
                result = asyncio.run(process_query_with_agents(multi_agent_system, user_input.strip(), st.session_state.history))
                answer = result.get("response_text", "")
                context = json.dumps(result.get("context_data", {}))
                sql_query = result.get("sql_query", None)
            elif use_enhanced and metadata_manager:
                answer, context, sql_query = asyncio.run(
                    process_enhanced_redshift_query(
                        user_input.strip(), redshift_engine, schema_info, metadata_manager, st.session_state.history
                    )
                )
            else:
                answer, context, sql_query = get_sql_answer(user_input.strip(), schema_info, st.session_state.history)
            save_chat(user_input.strip(), answer, context, sql_query, st.session_state.selected_session_id)
            st.session_state.history = get_chat_history(st.session_state.selected_session_id)
            st.rerun()

    for entry in reversed(st.session_state.history):
        with st.chat_message("user"):
            st.markdown(entry.question)
        with st.chat_message("assistant"):
            st.markdown("#### 💬 Response:")
            st.info(entry.answer)
            if entry.sql_query:
                st.markdown("#### 🔍 Generated SQL:")
                edited_sql = st.text_area(
                    "SQL Query (editable):", value=entry.sql_query, height=150, key=f"sql_edit_{entry.id}"
                )
                col1, col2 = st.columns([1, 1])
                with col1:
                    if st.button(f"▶️ Execute Query", key=f"exec_{entry.id}"):
                        with st.spinner("Executing query..."):
                            result_df = run_redshift_query(edited_sql, redshift_engine)
                            if result_df is not None:
                                st.markdown("#### 📊 Query Results:")
                                st.dataframe(result_df, use_container_width=True)
                with col2:
                    if st.button(f"📋 Copy SQL", key=f"copy_{entry.id}"):
                        st.code(edited_sql, language='sql')
            if entry == st.session_state.history[-1]:
                with st.expander("🔍 Follow-up Suggestions", expanded=False):
                    suggestions = generate_followup_suggestions_async(entry.question, entry.answer, entry.context)
                    for i, suggestion in enumerate(suggestions[:3]):
                        if st.button(f"💡 {suggestion}", key=f"sql_suggestion_{entry.id}_{i}"):
                            st.session_state.user_input = suggestion
                            st.rerun()

elif mode == "🔍 Enhanced Schema Explorer":
    st.subheader("🔍 Enhanced Schema Explorer")
    if not metadata_manager:
        st.warning("⚠️ Enhanced schema features require metadata manager initialization.")
        st.stop()
    display_enhanced_schema_explorer(metadata_manager)
    display_query_analytics(metadata_manager)

# Footer
st.markdown("---")
col1, col2, col3 = st.columns(3)
with col1:
    if st.button("🗑️ Clear Chat History"):
        if st.session_state.selected_session_id:
            db = SessionLocal()
            db.query(ChatHistory).filter_by(session_id=st.session_state.selected_session_id).delete()
            db.commit()
            db.close()
            st.session_state.history = []
            st.success("Chat history cleared!")
            st.rerun()

with col2:
    if st.button("📥 Export Chat"):
        if st.session_state.history:
            chat_export = [
                {
                    "timestamp": entry.timestamp.isoformat(),
                    "question": entry.question,
                    "answer": entry.answer,
                    "sql_query": entry.sql_query
                } for entry in st.session_state.history
            ]
            json_str = json.dumps(chat_export, indent=2)
            st.download_button(
                "📄 Download Chat History",
                data=json_str,
                file_name=f"chat_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                mime="application/json"
            )

with col3:
    st.metric("💬 Total Messages", len(st.session_state.history))
