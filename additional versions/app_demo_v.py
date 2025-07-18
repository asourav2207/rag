import streamlit as st
import os
import shutil
import pandas as pd
from datetime import datetime
from sentence_transformers import SentenceTransformer
import numpy as np
import faiss
import requests
import json
from sqlalchemy import create_engine, Column, Integer, String, Text, DateTime
from sqlalchemy.orm import declarative_base, sessionmaker
import base64
import hashlib
from io import BytesIO
from openpyxl import load_workbook
import concurrent.futures

# ---- GitHub Integration Config ----
# Store your GitHub token in Streamlit secrets or as an environment variable
GITHUB_TOKEN = st.secrets.get("github_token", os.environ.get("GITHUB_TOKEN", ""))
GITHUB_REPO = st.secrets.get("github_repo", "asourav2207/rag")  # e.g., "adityasourav/redshift-excels"
GITHUB_BRANCH = st.secrets.get("github_branch", "main")
GITHUB_EXCEL_PATH = st.secrets.get("github_excel_path", "excel_files")  # folder in repo

def github_headers():
    return {
        "Authorization": f"token {GITHUB_TOKEN}",
        "Accept": "application/vnd.github.v3+json"
    }

def list_github_excels():
    url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{GITHUB_EXCEL_PATH}?ref={GITHUB_BRANCH}"
    try:
        r = requests.get(url, headers=github_headers(), timeout=10)
        r.raise_for_status()
        files = r.json()
        return [f for f in files if f["name"].endswith(".xlsx")]
    except Exception as e:
        st.error(f"Error listing Excel files from GitHub: {e}")
        return []

def download_excel_from_github(file_info):
    try:
        r = requests.get(file_info["download_url"], timeout=10)
        r.raise_for_status()
        return pd.read_excel(r.content, engine="openpyxl")
    except Exception as e:
        st.error(f"Error downloading {file_info['name']} from GitHub: {e}")
        return None

def upload_excel_to_github(file, filename):
    url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{GITHUB_EXCEL_PATH}/{filename}"
    try:
        # Get SHA if file exists (for update)
        get_resp = requests.get(url, headers=github_headers(), timeout=10)
        sha = get_resp.json().get("sha") if get_resp.status_code == 200 else None

        file.seek(0)
        content = base64.b64encode(file.read()).decode()
        data = {
            "message": f"Add {filename}",
            "content": content,
            "branch": GITHUB_BRANCH,
        }
        if sha:
            data["sha"] = sha
        put_resp = requests.put(url, headers=github_headers(), data=json.dumps(data), timeout=15)
        if put_resp.status_code in (200, 201):
            return True
        else:
            st.error(f"GitHub upload failed for {filename}: {put_resp.json().get('message', put_resp.text)}")
            return False
    except Exception as e:
        st.error(f"Error uploading {filename} to GitHub: {e}")
        return False

# ---- Changelog Functions ----
def generate_changelog(existing_df, new_df):
    diffs = []
    max_len = max(len(existing_df), len(new_df))
    for i in range(max_len):
        if i >= len(existing_df):
            diffs.append(f"Row {i+1} added: {new_df.iloc[i].to_dict()}")
        elif i >= len(new_df):
            diffs.append(f"Row {i+1} removed: {existing_df.iloc[i].to_dict()}")
        else:
            old_row = existing_df.iloc[i].to_dict()
            new_row = new_df.iloc[i].to_dict()
            field_name = new_row.get("Field Name") or old_row.get("Field Name") or f"Row{i+1}"
            for col in new_row.keys():
                old_val = str(old_row.get(col, "")).strip()
                new_val = str(new_row.get(col, "")).strip()
                if old_val != new_val:
                    diffs.append(
                        f"Row {i+1} - Column '{col}' changed from '{old_val}' to '{new_val}' [Field Name: {field_name}]"
                    )
    return diffs

def inject_changelog_to_excel(uploaded_file, changelog_lines, user_id):
    uploaded_file.seek(0)
    wb = load_workbook(uploaded_file)
    if "Changelog" not in wb.sheetnames:
        ws = wb.create_sheet("Changelog")
    else:
        ws = wb["Changelog"]

    from datetime import datetime
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append(["Timestamp", "User", "Change Description"])
    for line in changelog_lines:
        ws.append([timestamp, user_id, line])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ---- SQLAlchemy ORM Setup ----
Base = declarative_base()
DB_PATH = "sqlite:///chat_history.db"

class ChatSession(Base):
    __tablename__ = 'chat_session'
    id = Column(Integer, primary_key=True)
    name = Column(String, default="Untitled Chat")
    created_at = Column(DateTime, default=datetime.now)

class ChatHistory(Base):
    __tablename__ = 'chat_history'
    id = Column(Integer, primary_key=True)
    question = Column(Text)
    answer = Column(Text)
    context = Column(Text)
    timestamp = Column(DateTime)
    feedback = Column(String)
    session_id = Column(Integer)

engine = create_engine(DB_PATH)
Base.metadata.create_all(engine)
SessionLocal = sessionmaker(bind=engine)

# ---- DB Functions ----
def save_chat(question, answer, context, session_id):
    db = SessionLocal()
    entry = ChatHistory(
        question=question,
        answer=answer,
        context=context,
        timestamp=datetime.now(),
        feedback=None,
        session_id=session_id,
    )
    db.add(entry)
    db.commit()
    db.close()

def get_chat_history(session_id):
    db = SessionLocal()
    entries = db.query(ChatHistory).filter_by(session_id=session_id).order_by(ChatHistory.timestamp).all()
    db.close()
    return entries

def clear_chat_history():
    db = SessionLocal()
    db.query(ChatHistory).delete()
    db.commit()
    db.close()

def download_chat_history():
    data_tuples = [(c.timestamp.strftime("%Y-%m-%d %H:%M:%S"), c.question, c.answer, c.context, c.feedback) for c in get_chat_history(st.session_state.selected_session_id)]
    df = pd.DataFrame(data_tuples, columns=["Timestamp", "Question", "Answer", "Context", "Feedback"])
    return df.to_csv(index=False).encode("utf-8")

# ---- Load and Prepare Data ----
@st.cache_resource(show_spinner=True)
def load_and_prepare():
    def process_row(row, df_columns, table_name, idx, row_docs, field_docs, row_metadata, field_metadata):
        row_summary = []
        for col in df_columns:
            val = str(row.get(col, '')).strip()
            if val:
                row_summary.append(f"{col.strip()}: {val}")
                field_docs.append(f"{col.strip()}: {val}\nTable: {table_name}")
                field_metadata.append({'table': table_name, 'column': col, 'row': idx})
        if row_summary:
            doc_text = "\n".join(row_summary) + f"\nTable: {table_name}, Row: {idx}"
            row_docs.append(doc_text)
            row_metadata.append({'table': table_name, 'row_index': idx})

    def process_excel_file(file_info, row_docs, field_docs, row_metadata, field_metadata):
        df = download_excel_from_github(file_info)
        if df is None:
            return
        table_name = os.path.splitext(file_info["name"])[0]
        for idx, row in df.iterrows():
            process_row(row, df.columns, table_name, idx, row_docs, field_docs, row_metadata, field_metadata)

    row_docs = []
    field_docs = []
    row_metadata = []
    field_metadata = []

    # --- Load Excel files from GitHub ---
    excel_files = list_github_excels()
    if not excel_files:
        st.warning("No Excel files found in the GitHub repository.")
    for file_info in excel_files:
        process_excel_file(file_info, row_docs, field_docs, row_metadata, field_metadata)

    if not row_docs or not field_docs:
        st.warning("No data found in the Excel files.")
        # Return empty objects to avoid downstream errors
        model = SentenceTransformer('all-MiniLM-L6-v2')
        dim = 384
        row_index = faiss.IndexFlatL2(dim)
        field_index = faiss.IndexFlatL2(dim)
        return model, row_index, [], [], field_index, [], []

    model = SentenceTransformer('all-MiniLM-L6-v2')
    row_embeddings = model.encode(row_docs, show_progress_bar=True)
    field_embeddings = model.encode(field_docs, show_progress_bar=True)

    dim = row_embeddings.shape[1]
    row_index = faiss.IndexFlatL2(dim)
    row_index.add(np.array(row_embeddings))

    field_index = faiss.IndexFlatL2(dim)
    field_index.add(np.array(field_embeddings))

    # Create a fast field name lookup for context
    field_lookup = {f"{fm['table']}.{fm['column']}": f"{fm['table']}.{fm['column']}: {fd}" for fm, fd in zip(field_metadata, field_docs)}

    return model, row_index, row_docs, row_metadata, field_index, field_docs, field_metadata, field_lookup

# ---- Query Handling ----
def retrieve_relevant_docs(query, model, row_index, row_docs, row_metadata, field_index, field_docs, field_metadata, field_lookup, k=5):
    # First, try exact match for field or element name in the docs
    query_clean = query.strip().lower()
    # Try to find an exact match in field_docs or row_docs
    for i, fd in enumerate(field_docs):
        # Extract field name from the doc (before ':')
        field_name = fd.split(':', 1)[0].strip().lower()
        if field_name == query_clean:
            # Return only the exact match row as context
            return [(field_docs[i], field_metadata[i])]
    for i, rd in enumerate(row_docs):
        # Try to match the field name in the row doc
        if query_clean in rd.lower().split(':', 1)[0]:
            return [(row_docs[i], row_metadata[i])]
    # Fallback: fuzzy/semantic search as before
    query_lower = query_clean
    for field, desc in field_lookup.items():
        if field.lower() in query_lower or query_lower in field.lower():
            return [(f"{field}: {desc}", {"field": field})]
    if not row_docs or not field_docs:
        return []
    query_emb = model.encode([query])
    dr, ir = row_index.search(np.array(query_emb), k)
    df, if_ = field_index.search(np.array(query_emb), k)
    row_results = [(row_docs[i], row_metadata[i]) for i, dist in zip(ir[0], dr[0]) if dist < 1.5]
    field_results = [(field_docs[i], field_metadata[i]) for i, dist in zip(if_[0], df[0]) if dist < 1.5]
    combined = dict(row_results + field_results)
    return list(combined.items())

def call_ollama(prompt):
    try:
        # Use 127.0.0.1 instead of localhost for better compatibility
        response = requests.post(
            "http://127.0.0.1:11434/api/generate",
            json={"model": "llama3.2", "prompt": prompt, "stream": False},
            timeout=60  # Increased timeout
        )
        response.raise_for_status()
        data = response.json()
        print("\n🟩 Ollama full JSON response:\n", json.dumps(data, indent=2))
        return data.get("response", "").strip()
    except requests.exceptions.RequestException as e:
        st.error(f"❌ Error contacting Ollama server: {e}\nDetails: {getattr(e, 'response', None)}")
        print(f"❌ Error contacting Ollama server: {e}\nDetails: {getattr(e, 'response', None)}")
        return f"❌ Error contacting Ollama server: {e}"
    except Exception as e:
        st.error(f"❌ Unexpected error: {e}")
        print(f"❌ Unexpected error: {e}")
        return f"❌ Unexpected error: {e}"

def answer_query(user_query, model, row_index, row_docs, row_metadata, field_index, field_docs, field_metadata, field_lookup):
    # Track last referenced field in session state
    ambiguous_terms = ["this field", "it", "the field", "that field"]
    user_query_lower = user_query.strip().lower()
    last_field = st.session_state.get("last_referenced_field")
    clarified_query = user_query
    # If the question is ambiguous and we have a last field, clarify it
    if any(term in user_query_lower for term in ambiguous_terms) and last_field:
        clarified_query = user_query + f" (referring to {last_field})"
        st.info(f"Clarified ambiguous question to: '{clarified_query}'")

    top_matches = retrieve_relevant_docs(clarified_query, model, row_index, row_docs, row_metadata, field_index, field_docs, field_metadata, field_lookup)
    # Limit context to top 2 matches or 2000 characters
    context_docs = [doc for doc, meta in top_matches][:2]
    context_str = "\n---\n".join(context_docs)
    max_context_len = 2000
    if len(context_str) > max_context_len:
        st.warning("Context is too large, truncating for faster LLM response.")
        context_str = context_str[:max_context_len]

    # Try to extract the referenced field from the top match for future follow-ups
    referenced_field = None
    if top_matches:
        meta = top_matches[0][1]
        if isinstance(meta, dict) and "field" in meta:
            referenced_field = meta["field"]
        else:
            import re
            m = re.search(r"([a-zA-Z0-9_]+)", context_docs[0])
            if m:
                referenced_field = m.group(1)
    if referenced_field:
        st.session_state["last_referenced_field"] = referenced_field

    # --- Add last 2 user+assistant turns as conversation history ---
    history = st.session_state.get("chat_history", [])
    conversation_history = ""
    if len(history) > 0:
        # Get last 2 turns (user+assistant)
        last_turns = history[-2:]
        for entry in last_turns:
            conversation_history += f"User: {entry.question}\nAssistant: {entry.answer}\n"

    prompt = f"""You are a Redshift assistant. Here is the recent conversation history (if any):\n{conversation_history}\n\nUse the following context to explain the answer in detail if necessary. If the user requests SQL, generate the SQL snippet.\n\nContext:\n{context_str}\n\nUser query:\n{clarified_query}\n\nReply clearly. Include SQL if asked. Keep responses conversational but thorough."""
    answer = call_ollama(prompt)
    save_chat(user_query, answer, context_str, st.session_state.selected_session_id)
    return answer.strip(), top_matches

# ---- Follow-up Suggestions (LLM-based, improved prompt and answer on click) ----
def generate_followup_suggestions_async(user_prompt, system_response, context, timeout=20):
    """
    Use a thread pool to generate follow-up suggestions asynchronously, with a timeout.
    Suggestions are phrased as direct questions the user might ask next.
    """
    def _llm_suggest():
        prompt = f"""
Given the following user question, system answer, and context, generate 2-3 direct, concise follow-up questions that a user might naturally ask next. Phrase each as a question, not as a statement or suggestion. If the context is code or SQL, offer to explain or generate related code. If the answer is a definition, suggest related fields or deeper exploration. Do not use phrases like 'the user might want...'.

User question: {user_prompt}
System answer: {system_response}
Context: {context}

List the follow-up questions as bullet points, phrased as questions.
"""
        suggestions = call_ollama(prompt)
        # Parse suggestions as bullet points/questions
        return [line.lstrip('-• ').strip().rstrip('?') + '?' for line in suggestions.split('\n') if line.strip() and (line.startswith('-') or line.startswith('•'))]

    with concurrent.futures.ThreadPoolExecutor() as executor:
        future = executor.submit(_llm_suggest)
        try:
            return future.result(timeout=timeout)
        except concurrent.futures.TimeoutError:
            return ["(Follow-up suggestions timed out. Please try again.)"]

# ---- Streamlit UI ----
st.set_page_config(page_title="Redshift RAG Chatbot", layout="wide")
st.title("🧑‍💻 Redshift Documentation Assistant")

# Sidebar for chat session management
st.sidebar.title("💬 Chat Sessions")
with st.sidebar.form("new_session"):
    new_chat_name = st.text_input("New chat name")
    if st.form_submit_button("Start New Chat") and new_chat_name:
        db = SessionLocal()
        new_session = ChatSession(name=new_chat_name)
        db.add(new_session)
        db.commit()
        st.session_state.selected_session_id = new_session.id
        st.session_state.chat_history = []
        db.close()
        st.rerun()

db = SessionLocal()
sessions = db.query(ChatSession).order_by(ChatSession.created_at.desc()).all()
db.close()
session_options = {s.name: s.id for s in sessions}

if 'selected_session_id' not in st.session_state:
    if sessions:
        st.session_state.selected_session_id = sessions[0].id
    else:
        # Create a default session if none exist
        db = SessionLocal()
        default_session = ChatSession(name="Default Chat")
        db.add(default_session)
        db.commit()
        st.session_state.selected_session_id = default_session.id
        db.close()

selected_chat_name = None
if session_options:
    selected_chat_name = st.sidebar.selectbox("Select Chat", list(session_options.keys()), index=0)
    st.session_state.selected_session_id = session_options[selected_chat_name]

if 'chat_history' not in st.session_state or st.session_state.get("last_session_id") != st.session_state.selected_session_id:
    st.session_state.chat_history = get_chat_history(st.session_state.selected_session_id)
    st.session_state.last_session_id = st.session_state.selected_session_id

# ---- Upload Excel files to GitHub ----
enable_changelog = st.toggle("Enable Changelog Tracking", value=True)
uploaded_files = st.file_uploader("Upload Redshift Excel files (uploaded to GitHub repo)", type=["xlsx"], accept_multiple_files=True)
if uploaded_files:
    for file in uploaded_files:
        file.seek(0)
        existing_df = None
        existing_file_info = [f for f in list_github_excels() if f["name"] == file.name]
        if existing_file_info:
            existing_df = download_excel_from_github(existing_file_info[0])

        uploaded_df = pd.read_excel(file, engine="openpyxl")
        changelog = []

        if enable_changelog and existing_df is not None:
            changelog = generate_changelog(existing_df, uploaded_df)

        user_id = st.text_input("Enter your GitHub user ID", key=f"user_id_{file.name}") if enable_changelog else "unknown"

        if enable_changelog and changelog:
            file_with_changelog = inject_changelog_to_excel(file, changelog, user_id)
            file.seek(0)
            file_with_changelog.seek(0)
            file_to_upload = file_with_changelog
        else:
            file.seek(0)
            file_to_upload = file

        if upload_excel_to_github(file_to_upload, file.name):
            st.success(f"✅ {file.name} uploaded to GitHub with changelog." if changelog else f"✅ {file.name} uploaded to GitHub.")
            # --- Smart Suggestions Section ---
            with st.expander(f"💡 Smart Suggestions based on {file.name}"):
                columns = uploaded_df.columns.str.lower()
                suggestions = []
                if "member_id" in columns and "claim_id" in columns:
                    suggestions.append("How do I join member and claim tables?")
                if "amount" in columns or "payment" in columns:
                    suggestions.append("What logic is used to calculate payment amount?")
                if any("date" in col for col in columns):
                    suggestions.append("Are all date fields in the same format?")
                # Add default general-purpose fallback
                if len(suggestions) < 3:
                    extra = [
                        "What are the key joins among uploaded tables?",
                        "Can you generate sample Redshift SQL using this file?",
                        "Which columns have missing values or nulls?",
                    ]
                    for s in extra:
                        if len(suggestions) >= 3:
                            break
                        suggestions.append(s)
                # Display up to 3 buttons
                for i, q in enumerate(suggestions[:3]):
                    if st.button(f"💬 {q}", key=f"suggested_{file.name}_{i}"):
                        st.session_state["user_input"] = q
                        st.rerun()
        else:
            st.error(f"❌ Failed to upload {file.name} to GitHub.")
    st.info("Please refresh the page after upload.")

with st.spinner("Loading and embedding documents..."):
    model, row_index, row_docs_list, row_metadata_list, field_index, field_docs_list, field_metadata_list, field_lookup = load_and_prepare()

user_input = st.text_input(
    f"Ask a question about your Redshift columns or table joins (Session: {selected_chat_name}):",
    placeholder="E.g., What does user_id mean? How do I join claims and members?",
    key="user_input"
)

if st.button("Ask") and user_input.strip():
    with st.spinner("🧠 Thinking..."):
        answer, matches = answer_query(
            user_input.strip(),
            model,
            row_index, row_docs_list, row_metadata_list,
            field_index, field_docs_list, field_metadata_list, field_lookup
        )
        st.session_state.chat_history.append(ChatHistory(
            question=user_input.strip(),
            answer=answer,
            context="\n\n".join([doc for doc, _ in matches]),
            timestamp=datetime.now(),
            session_id=st.session_state.selected_session_id
        ))
        # Store last question and answer for follow-up suggestions
        st.session_state["last_question"] = user_input.strip()
        st.session_state["last_answer"] = answer

# ---- Handle Pending Follow-up (before UI rendering) ----
if st.session_state.get("pending_followup"):
    q = st.session_state.pop("pending_followup")
    with st.spinner("🧠 Generating answer for follow-up..."):
        answer, matches = answer_query(
            q,
            model,
            row_index, row_docs_list, row_metadata_list,
            field_index, field_docs_list, field_metadata_list, field_lookup
        )
        st.session_state.chat_history.append(ChatHistory(
            question=q,
            answer=answer,
            context="\n\n".join([doc for doc, _ in matches]),
            timestamp=datetime.now(),
            session_id=st.session_state.selected_session_id
        ))
        st.session_state["last_question"] = q
        st.session_state["last_answer"] = answer
        st.rerun()

# ---- Show Chat History ----
for entry in st.session_state.chat_history[::-1]:
    with st.chat_message("user"):
        st.markdown(entry.question)
    with st.chat_message("assistant"):
        st.markdown("#### 💬 Answer:")
        st.success(entry.answer if entry.answer else "Sorry, no answer generated.")
        # --- Follow-up Suggestions Section ---
        if entry.question == st.session_state.get("last_question"):
            with st.expander("🔎 Follow-up Suggestions"):
                suggestions = generate_followup_suggestions_async(
                    entry.question,
                    entry.answer,
                    entry.context
                )
                for i, q in enumerate(suggestions):
                    if st.button(f"👉 {q}", key=f"followup_{entry.id}_{i}"):
                        st.session_state["pending_followup"] = q
                        st.rerun()
        with st.expander("📄 Show Retrieved Context"):
            for i, doc in enumerate(entry.context.strip().split("\n\n")):
                st.markdown(f"**Context {i+1}**:")
                st.code(doc)

# ---- Tools ----
st.download_button("📥 Download Chat History", data=download_chat_history(), file_name="chat_history.csv")
if st.button("🗑️ Clear All History"):
    clear_chat_history()
    st.session_state.chat_history = []
    st.success("Chat history cleared. Please refresh.")
